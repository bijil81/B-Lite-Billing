"""
accounting.py  Ã¢â‚¬â€œ  BOBY'S Salon : Full Accounting
Income / Expense / Profit comparison
Weekly / Monthly / Yearly reports + Print
FIXES:
  - Fix R4a: _get_revenue() file read try/except Ã¢â‚¬â€ crash prevention
  - Fix R4b: _get_revenue_by_day() file read try/except Ã¢â‚¬â€ crash prevention
  - Fix R4c: _get_expenses_by_day() try/except Ã¢â‚¬â€ crash prevention
  - Fix R4d: _calc_summary() wrapped with error guard (_calc_summary_inner)
  - Fix R4e: _calc_comparison() wrapped with error guard (_calc_comparison_inner)
  - Fix R4f: _gen_detailed() wrapped with error guard (_gen_detailed_inner)
  - Fix R4g: _export_excel() cross-platform file open + outer try/except
  - Fix R4h: _build_report_text() try/except Ã¢â‚¬â€ returns error string on fail
  - Fix R4i: refresh() try/except Ã¢â‚¬â€ crash prevention
"""
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date, timedelta
import csv, os
from collections import defaultdict
from branding import get_company_name
from utils import (C, safe_float, fmt_currency, F_REPORT,
                   DATA_DIR, today_str, month_str, app_log)
from expenses import get_expenses
from ui_theme import ModernButton


# Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
#  DATA HELPERS
# Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
def _get_revenue(from_d: str, to_d: str) -> float:
    total = 0.0
    if not os.path.exists(F_REPORT): return 0.0
    try:
        with open(F_REPORT, "r", encoding="utf-8") as f:
            r   = csv.reader(f)
            hdr = next(r, None)
            ti  = 5 if (hdr and len(hdr) >= 6) else 3
            for row in r:
                if row and len(row) > ti:
                    dt = row[0][:10]
                    if from_d <= dt <= to_d:
                        total += safe_float(row[ti])
    except Exception as e:
        app_log(f"[_get_revenue] {e}")
    return total


def _get_revenue_by_day(from_d: str, to_d: str) -> dict:
    data = defaultdict(float)
    if not os.path.exists(F_REPORT): return data
    try:
        with open(F_REPORT, "r", encoding="utf-8") as f:
            r   = csv.reader(f)
            hdr = next(r, None)
            ti  = 5 if (hdr and len(hdr) >= 6) else 3
            for row in r:
                if row and len(row) > ti:
                    dt = row[0][:10]
                    if from_d <= dt <= to_d:
                        data[dt] += safe_float(row[ti])
    except Exception as e:
        app_log(f"[_get_revenue_by_day] {e}")
    return data


def _get_expenses_by_day(from_d: str, to_d: str) -> dict:
    data = defaultdict(float)
    try:
        for e in get_expenses():
            dt = e.get("date", "")[:10]
            if from_d <= dt <= to_d:
                data[dt] += safe_float(e.get("amount", 0))
    except Exception as e:
        app_log(f"[_get_expenses_by_day] {e}")
    return data


def _period_dates(period: str):
    """Returns (from_d, to_d, label) for a period string."""
    today = date.today()

    if period == "today":
        d = today.strftime("%Y-%m-%d")
        return d, d, f"Today ({d})"

    elif period == "yesterday":
        d = (today - timedelta(days=1)).strftime("%Y-%m-%d")
        return d, d, f"Yesterday ({d})"

    elif period == "this_week":
        monday = today - timedelta(days=today.weekday())
        return (monday.strftime("%Y-%m-%d"),
                today.strftime("%Y-%m-%d"),
                f"This Week ({monday.strftime('%d %b')} Ã¢â‚¬â€œ {today.strftime('%d %b')})")

    elif period == "last_week":
        monday = today - timedelta(days=today.weekday()+7)
        sunday = monday + timedelta(days=6)
        return (monday.strftime("%Y-%m-%d"),
                sunday.strftime("%Y-%m-%d"),
                f"Last Week ({monday.strftime('%d %b')} Ã¢â‚¬â€œ {sunday.strftime('%d %b')})")

    elif period == "this_month":
        fd = today.strftime("%Y-%m-01")
        return fd, today.strftime("%Y-%m-%d"), \
               f"This Month ({today.strftime('%B %Y')})"

    elif period == "last_month":
        first = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last  = today.replace(day=1) - timedelta(days=1)
        return (first.strftime("%Y-%m-%d"), last.strftime("%Y-%m-%d"),
                f"Last Month ({first.strftime('%B %Y')})")

    elif period == "last_3_months":
        fd = (today - timedelta(days=90)).strftime("%Y-%m-%d")
        return fd, today.strftime("%Y-%m-%d"), "Last 3 Months"

    elif period == "last_6_months":
        fd = (today - timedelta(days=180)).strftime("%Y-%m-%d")
        return fd, today.strftime("%Y-%m-%d"), "Last 6 Months"

    elif period == "this_year":
        fd = today.strftime("%Y-01-01")
        return fd, today.strftime("%Y-%m-%d"), \
               f"This Year ({today.year})"

    elif period == "last_year":
        y  = today.year - 1
        return f"{y}-01-01", f"{y}-12-31", f"Last Year ({y})"

    elif period == "all":
        return "2000-01-01", today.strftime("%Y-%m-%d"), "All Time"

    return today.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d"), "Today"


# Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
#  ACCOUNTING FRAME
# Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
class AccountingFrame(tk.Frame):

    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app = app
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["card"], pady=10)
        hdr.pack(fill=tk.X)
        left_f = tk.Frame(hdr, bg=C["card"])
        left_f.pack(side=tk.LEFT, padx=20)
        tk.Label(left_f, text="Full Accounting",
                 font=("Arial", 15, "bold"),
                 bg=C["card"], fg=C["text"]).pack(anchor="w")
        tk.Label(left_f, text="Revenue, expenses & profit analysis",
                 font=("Arial", 10), bg=C["card"], fg=C["muted"]).pack(anchor="w")
        ModernButton(hdr, text="Print",
                     command=self._print_report,
                     color=C["blue"], hover_color="#154360",
                     width=90, height=32, radius=8,
                     font=("Arial",10,"bold"),
                     ).pack(side=tk.RIGHT, padx=5, pady=6)
        ModernButton(hdr, text="Excel",
                     command=self._export_excel,
                     color=C["green"], hover_color="#1a7a45",
                     width=90, height=32, radius=8,
                     font=("Arial",10,"bold"),
                     ).pack(side=tk.RIGHT, padx=5, pady=6)
        tk.Frame(self, bg=C["teal"], height=2).pack(fill=tk.X)

        nb = ttk.Notebook(self)
        nb.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        t1 = tk.Frame(nb, bg=C["bg"])
        t2 = tk.Frame(nb, bg=C["bg"])
        t3 = tk.Frame(nb, bg=C["bg"])
        nb.add(t1, text="Summary")
        nb.add(t2, text="Comparison")
        nb.add(t3, text="Detailed Report")


        self._build_summary(t1)
        self._build_comparison(t2)
        self._build_detailed(t3)

    # Ã¢â€â‚¬Ã¢â€â‚¬ SUMMARY TAB Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    def _build_summary(self, parent):
        # Period selector
        ctrl = tk.Frame(parent, bg=C["bg"], pady=12)
        ctrl.pack(fill=tk.X, padx=20)

        tk.Label(ctrl, text="Period:", bg=C["bg"],
                 fg=C["muted"], font=("Arial", 11, "bold")).pack(side=tk.LEFT, padx=(0,10))

        self.period_var = tk.StringVar(value="this_month")
        periods = [
            ("Today",        "today"),
            ("Yesterday",    "yesterday"),
            ("This Week",    "this_week"),
            ("Last Week",    "last_week"),
            ("This Month",   "this_month"),
            ("Last Month",   "last_month"),
            ("3 Months",     "last_3_months"),
            ("6 Months",     "last_6_months"),
            ("This Year",    "this_year"),
            ("Last Year",    "last_year"),
            ("All Time",     "all"),
        ]
        pcb = ttk.Combobox(ctrl, textvariable=self.period_var,
                            values=[p[0] for p in periods],
                            state="readonly", font=("Arial", 11), width=14)
        pcb.pack(side=tk.LEFT, padx=(0,10))
        self._period_map = {p[0]: p[1] for p in periods}

        ModernButton(ctrl, text="Calculate",
                     command=self._calc_summary,
                     color=C["teal"], hover_color=C["blue"],
                     width=120, height=34, radius=8,
                     font=("Arial",10,"bold"),
                     ).pack(side=tk.LEFT)

        self.period_lbl = tk.Label(ctrl, text="",
                                    bg=C["bg"], fg=C["muted"],
                                    font=("Arial", 12))
        self.period_lbl.pack(side=tk.LEFT, padx=15)

        # Summary cards
        self.sum_cards = tk.Frame(parent, bg=C["bg"])
        self.sum_cards.pack(fill=tk.X, padx=20, pady=(0, 12))

        # Detailed breakdown table
        cols = ("Category", "Amount")
        self.sum_tree = ttk.Treeview(parent, columns=cols,
                                      show="headings", height=18)
        self.sum_tree.heading("Category", text="Category")
        self.sum_tree.heading("Amount",   text="Amount")
        self.sum_tree.column("Category",  width=400)
        self.sum_tree.column("Amount",    width=200, anchor="e")
        self.sum_tree.pack(fill=tk.BOTH, expand=True, padx=20)

        # Expense breakdown
        _etf_o = tk.Frame(parent, bg=C["card"])
        _etf_o.pack(fill=tk.X, padx=20, pady=8)
        _etfh = tk.Frame(_etf_o, bg=C["sidebar"], padx=12, pady=6)
        _etfh.pack(fill=tk.X)
        tk.Label(_etfh, text="Expense Breakdown", font=("Arial",11,"bold"),
                 bg=C["sidebar"], fg=C["text"]).pack(side=tk.LEFT)
        tk.Frame(_etf_o, bg=C["orange"], height=2).pack(fill=tk.X)
        self._exp_tree_frame = tk.Frame(_etf_o, bg=C["bg"], padx=10, pady=8)
        self._exp_tree_frame.pack(fill=tk.X)

        cols2 = ("Category", "Count", "Total")
        self.exp_tree = ttk.Treeview(self._exp_tree_frame, columns=cols2,
                                      show="headings", height=6)
        for col, w in zip(cols2, [250, 80, 150]):
            self.exp_tree.heading(col, text=col)
            self.exp_tree.column(col, width=w,
                                  anchor="e" if col!="Category" else "w")
        self.exp_tree.pack(fill=tk.X)

        self._calc_summary()

    def _calc_summary(self):
        try:
            self._calc_summary_inner()
        except Exception as e:
            app_log(f"[_calc_summary] {e}")

    def _calc_summary_inner(self):
        period_name = self.period_var.get() or "This Month"
        period_key  = self._period_map.get(period_name, "this_month")
        from_d, to_d, label = _period_dates(period_key)
        self.period_lbl.config(text=label)

        # Revenue
        revenue  = _get_revenue(from_d, to_d)

        # Expenses
        expenses  = get_expenses()
        exp_total = sum(safe_float(e.get("amount", 0))
                        for e in expenses
                        if from_d <= e.get("date","")[:10] <= to_d)

        # Expense by category
        exp_by_cat = defaultdict(list)
        for e in expenses:
            if from_d <= e.get("date","")[:10] <= to_d:
                exp_by_cat[e.get("category","Other")].append(
                    safe_float(e.get("amount",0)))

        profit     = revenue - exp_total
        margin_pct = (profit / revenue * 100) if revenue > 0 else 0

        # Bills count
        bill_count = 0
        if os.path.exists(F_REPORT):
            with open(F_REPORT, "r", encoding="utf-8") as f:
                r   = csv.reader(f)
                hdr = next(r, None)
                for row in r:
                    if row and len(row) > 0 and from_d <= row[0][:10] <= to_d:
                        bill_count += 1

        avg_bill = revenue / bill_count if bill_count else 0

        # Update cards
        for w in self.sum_cards.winfo_children(): w.destroy()
        for lbl, val, col in [
            ("Revenue",    fmt_currency(revenue),  C["teal"]),
            ("Expenses",   fmt_currency(exp_total),C["red"]),
            ("Net Profit", fmt_currency(profit),
             C["green"] if profit >= 0 else C["red"]),
            ("Margin",     f"{margin_pct:.1f}%",   C["purple"]),
            ("Bills",      str(bill_count),         C["blue"]),
            ("Avg Bill",   fmt_currency(avg_bill),  C["orange"]),
        ]:
            card = tk.Frame(self.sum_cards, bg=col, padx=14, pady=10)
            card.pack(side=tk.LEFT, padx=(0,6))
            tk.Label(card, text=val, font=("Arial", 12, "bold"),
                     bg=col, fg="white").pack()
            tk.Label(card, text=lbl, font=("Arial", 10),
                     bg=col, fg="white").pack()

        # Summary tree
        for i in self.sum_tree.get_children(): self.sum_tree.delete(i)
        rows = [
            ("INCOME",                  ""),
            ("  Total Revenue",              fmt_currency(revenue)),
            ("  Number of Bills",            str(bill_count)),
            ("  Average Bill Value",         fmt_currency(avg_bill)),
            ("EXPENSES",                ""),
            ("  Total Expenses",             fmt_currency(exp_total)),
            ("PROFIT / LOSS",           ""),
            ("  Gross Profit",               fmt_currency(profit)),
            ("  Profit Margin",              f"{margin_pct:.1f}%"),
            ("  Revenue per Day",
             fmt_currency(revenue / max(1, (
                 date.fromisoformat(to_d) -
                 date.fromisoformat(from_d)).days + 1))),
        ]
        for cat, amt in rows:
            bold = cat in {"INCOME", "EXPENSES", "PROFIT / LOSS"}
            self.sum_tree.insert("", tk.END, values=(cat, amt),
                                  tags=("bold",) if bold else ())
        self.sum_tree.tag_configure("bold",
                                     font=("Arial", 12, "bold"),
                                     foreground=C["accent"])

        # Expense breakdown
        for i in self.exp_tree.get_children(): self.exp_tree.delete(i)
        for cat, amts in sorted(exp_by_cat.items(),
                                  key=lambda x: sum(x[1]), reverse=True):
            self.exp_tree.insert("", tk.END, values=(
                cat, len(amts), fmt_currency(sum(amts))))

        # Store for print/export
        self._last_summary = {
            "label": label, "from_d": from_d, "to_d": to_d,
            "revenue": revenue, "expenses": exp_total,
            "profit": profit, "margin": margin_pct,
            "bill_count": bill_count, "avg_bill": avg_bill,
            "exp_by_cat": dict(exp_by_cat),
        }

    # Ã¢â€â‚¬Ã¢â€â‚¬ COMPARISON TAB Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    def _build_comparison(self, parent):
        ctrl = tk.Frame(parent, bg=C["bg"], pady=12)
        ctrl.pack(fill=tk.X, padx=20)

        tk.Label(ctrl, text="Compare:",
                 bg=C["bg"], fg=C["muted"],
                 font=("Arial", 11, "bold")).pack(side=tk.LEFT, padx=(0,8))

        self.cmp_type = tk.StringVar(value="weekly")
        for txt, val in [("Weekly (last 8 weeks)", "weekly"),
                          ("Monthly (last 12 months)", "monthly"),
                          ("Yearly", "yearly")]:
            tk.Radiobutton(ctrl, text=txt,
                           variable=self.cmp_type, value=val,
                           bg=C["bg"], fg=C["text"],
                           selectcolor=C["input"],
                           font=("Arial", 12),
                           command=self._calc_comparison,
                           cursor="hand2").pack(side=tk.LEFT, padx=8)

        # Comparison table
        self.cmp_tree = ttk.Treeview(parent,
                                      columns=("Period","Revenue","Expenses","Profit","Bills"),
                                      show="headings", height=15)
        for col, w in zip(("Period","Revenue","Expenses","Profit","Bills"),
                           (160, 130, 130, 130, 80)):
            self.cmp_tree.heading(col, text=col)
            self.cmp_tree.column(col, width=w,
                                  anchor="e" if col!="Period" else "w")
        vsb = ttk.Scrollbar(parent, orient="vertical",
                             command=self.cmp_tree.yview)
        self.cmp_tree.configure(yscrollcommand=vsb.set)
        self.cmp_tree.pack(fill=tk.BOTH, expand=True,
                            padx=20, side=tk.LEFT)
        vsb.pack(side=tk.LEFT, fill=tk.Y, pady=(0,8))

        self._calc_comparison()

    def _calc_comparison(self):
        try:
            self._calc_comparison_inner()
        except Exception as e:
            app_log(f"[_calc_comparison] {e}")

    def _calc_comparison_inner(self):
        for i in self.cmp_tree.get_children(): self.cmp_tree.delete(i)
        ctype = self.cmp_type.get()
        today = date.today()
        periods = []

        if ctype == "weekly":
            for w in range(7, -1, -1):
                monday = today - timedelta(days=today.weekday() + 7*w)
                sunday = monday + timedelta(days=6)
                label  = f"W{monday.strftime('%d %b')}"
                periods.append((label,
                                  monday.strftime("%Y-%m-%d"),
                                  min(sunday, today).strftime("%Y-%m-%d")))

        elif ctype == "monthly":
            for m in range(11, -1, -1):
                d = (today.replace(day=1) - timedelta(days=1)) if m > 0 else today
                # Go back m months
                y, mo = today.year, today.month - m
                while mo <= 0: mo += 12; y -= 1
                first = date(y, mo, 1)
                if mo == 12: last = date(y+1,1,1) - timedelta(days=1)
                else:        last = date(y, mo+1,1) - timedelta(days=1)
                last = min(last, today)
                label = first.strftime("%b %Y")
                periods.append((label,
                                  first.strftime("%Y-%m-%d"),
                                  last.strftime("%Y-%m-%d")))

        elif ctype == "yearly":
            for y in range(today.year - 4, today.year + 1):
                periods.append((str(y), f"{y}-01-01",
                                  min(f"{y}-12-31",
                                      today.strftime("%Y-%m-%d"))))

        for label, fd, td in periods:
            rev  = _get_revenue(fd, td)
            exps = get_expenses()
            exp  = sum(safe_float(e.get("amount",0))
                       for e in exps if fd <= e.get("date","")[:10] <= td)
            profit = rev - exp
            bills  = 0
            if os.path.exists(F_REPORT):
                with open(F_REPORT,"r",encoding="utf-8") as f:
                    r = csv.reader(f)
                    next(r, None)
                    for row in r:
                        if row and fd <= row[0][:10] <= td:
                            bills += 1

            tag = "profit" if profit >= 0 else "loss"
            self.cmp_tree.insert("", tk.END, values=(
                label,
                fmt_currency(rev),
                fmt_currency(exp),
                fmt_currency(profit),
                bills,
            ), tags=(tag,))

        self.cmp_tree.tag_configure("profit", foreground="#27ae60")
        self.cmp_tree.tag_configure("loss",   foreground="#e74c3c")

    # Ã¢â€â‚¬Ã¢â€â‚¬ DETAILED REPORT TAB Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    def _build_detailed(self, parent):
        ctrl = tk.Frame(parent, bg=C["bg"], pady=10)
        ctrl.pack(fill=tk.X, padx=20)

        tk.Label(ctrl, text="From:", bg=C["bg"],
                 fg=C["muted"], font=("Arial", 12)).pack(side=tk.LEFT, padx=(0,4))
        self.det_from = tk.Entry(ctrl, font=("Arial", 12),
                                  bg=C["input"], fg=C["text"],
                                  bd=0, width=13, insertbackground=C["accent"])
        self.det_from.pack(side=tk.LEFT, ipady=5, padx=(0,10))
        self.det_from.insert(0, date.today().strftime("%Y-%m-01"))

        tk.Label(ctrl, text="To:", bg=C["bg"],
                 fg=C["muted"], font=("Arial", 12)).pack(side=tk.LEFT, padx=(0,4))
        self.det_to = tk.Entry(ctrl, font=("Arial", 12),
                                bg=C["input"], fg=C["text"],
                                bd=0, width=13, insertbackground=C["accent"])
        self.det_to.pack(side=tk.LEFT, ipady=5, padx=(0,10))
        self.det_to.insert(0, today_str())

        ModernButton(ctrl, text="Generate",
                     command=self._gen_detailed,
                     color=C["teal"], hover_color=C["blue"],
                     width=110, height=32, radius=8,
                     font=("Arial",10,"bold"),
                     ).pack(side=tk.LEFT)

        # Day-wise table
        cols = ("Date", "Revenue", "Expenses", "Profit", "Bills")
        self.det_tree = ttk.Treeview(parent, columns=cols,
                                      show="headings", height=20)
        for col, w in zip(cols, (130, 130, 130, 130, 80)):
            self.det_tree.heading(col, text=col)
            self.det_tree.column(col, width=w,
                                  anchor="e" if col!="Date" else "w")
        vsb = ttk.Scrollbar(parent, orient="vertical",
                             command=self.det_tree.yview)
        self.det_tree.configure(yscrollcommand=vsb.set)
        self.det_tree.pack(fill=tk.BOTH, expand=True,
                            padx=20, side=tk.LEFT)
        vsb.pack(side=tk.LEFT, fill=tk.Y, pady=(0,8))

    def _gen_detailed(self):
        try:
            self._gen_detailed_inner()
        except Exception as e:
            messagebox.showerror("Error", f"Could not generate report: {e}")

    def _gen_detailed_inner(self):
        for i in self.det_tree.get_children(): self.det_tree.delete(i)
        fd  = self.det_from.get().strip()
        td  = self.det_to.get().strip()
        rev = _get_revenue_by_day(fd, td)
        exp = _get_expenses_by_day(fd, td)

        all_dates = sorted(set(list(rev.keys()) + list(exp.keys())))
        tot_rev = tot_exp = tot_bills = 0

        for d in all_dates:
            r  = rev.get(d, 0.0)
            e  = exp.get(d, 0.0)
            p  = r - e
            b  = 0
            if os.path.exists(F_REPORT):
                with open(F_REPORT,"r",encoding="utf-8") as f:
                    cr = csv.reader(f)
                    next(cr, None)
                    for row in cr:
                        if row and row[0][:10] == d:
                            b += 1
            tot_rev += r; tot_exp += e; tot_bills += b
            tag = "profit" if p >= 0 else "loss"
            self.det_tree.insert("", tk.END, values=(
                d, fmt_currency(r), fmt_currency(e),
                fmt_currency(p), b), tags=(tag,))

        # Total row
        self.det_tree.insert("", tk.END,
                              values=("TOTAL",
                                       fmt_currency(tot_rev),
                                       fmt_currency(tot_exp),
                                       fmt_currency(tot_rev-tot_exp),
                                       tot_bills),
                              tags=("total",))
        self.det_tree.tag_configure("profit", foreground="#27ae60")
        self.det_tree.tag_configure("loss",   foreground="#e74c3c")
        self.det_tree.tag_configure("total",
                                     font=("Arial",12,"bold"),
                                     foreground=C["gold"])

    # Ã¢â€â‚¬Ã¢â€â‚¬ PRINT Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    def _print_report(self):
        if not hasattr(self, "_last_summary"):
            self._calc_summary()
        s = self._last_summary
        text  = self._build_report_text(s)
        try:
            import win32print
            pn   = win32print.GetDefaultPrinter()
            hprn = win32print.OpenPrinter(pn)
            try:
                win32print.StartDocPrinter(hprn,1,("Accounting Report",None,"RAW"))
                win32print.StartPagePrinter(hprn)
                win32print.WritePrinter(hprn, text.encode("utf-8"))
                win32print.EndPagePrinter(hprn)
                win32print.EndDocPrinter(hprn)
                messagebox.showinfo("Printed", f"Report printed to: {pn}")
            finally:
                win32print.ClosePrinter(hprn)
        except Exception as e:
            messagebox.showerror("Print Error", f"Could not print:\n{e}")

    def _build_report_text(self, s: dict) -> str:
        try:
            from salon_settings import get_settings
            cfg  = get_settings()
            W    = 52
            now  = datetime.now().strftime("%d-%m-%Y %I:%M %p")
            def L(t=""): return t+"\n"
            sname = cfg.get('salon_name', get_company_name())
            r  = L(f"{sname:^{W}}")
            r += L(f"{'ACCOUNTING REPORT':^{W}}")
            r += L(f"{s['label']:^{W}}")
            r += L(f"Generated: {now:^{W-11}}")
            r += L("="*W)
            r += L(f"{'INCOME SUMMARY':^{W}}")
            r += L("-"*W)
            r += L(f"  Total Revenue{s['revenue']:>{W-15}.2f}")
            r += L(f"  Total Bills{s['bill_count']:>{W-13}}")
            r += L(f"  Average Bill{s['avg_bill']:>{W-14}.2f}")
            r += L("-"*W)
            r += L(f"  Total Expenses{s['expenses']:>{W-16}.2f}")
            r += L("="*W)
            r += L(f"  NET PROFIT{s['profit']:>{W-12}.2f}")
            r += L(f"  Profit Margin{s['margin']:>{W-14}.1f}%")
            r += L("="*W)
            r += L(f"\n{'EXPENSE BREAKDOWN':^{W}}")
            r += L("-"*W)
            for cat, amts in sorted(s["exp_by_cat"].items(),
                                      key=lambda x: sum(x[1]), reverse=True):
                total = sum(amts)
                r += L(f"  {cat:<30} {total:>10.2f}")
            r += L("="*W)
            r += L(f"\n{'Thank You':^{W}}")
            return r
        except Exception as e:
            app_log(f"[_build_report_text] {e}")
            return "Error building report text."

    # Ã¢â€â‚¬Ã¢â€â‚¬ EXCEL EXPORT Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    def _export_excel(self):
        if not hasattr(self, "_last_summary"):
            self._calc_summary()
        s = self._last_summary
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()

            # Summary sheet
            ws = wb.active
            ws.title = "Summary"
            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill("solid", fgColor="1a1a2e")

            ws.append([f"{get_company_name()} - Accounting Report"])
            ws.append([s["label"]])
            ws.append([])
            for row in [
                ["Revenue",    s["revenue"]],
                ["Expenses",   s["expenses"]],
                ["Net Profit", s["profit"]],
                ["Margin %",   round(s["margin"],2)],
                ["Bills",      s["bill_count"]],
                ["Avg Bill",   round(s["avg_bill"],2)],
            ]:
                ws.append(row)

            out = os.path.join(DATA_DIR,
                                f"accounting_{today_str()}.xlsx")
            wb.save(out)
            messagebox.showinfo("Exported", f"Excel saved:\n{out}")
            try:
                from utils import open_file_cross_platform
                open_file_cross_platform(out)
            except Exception:
                pass
        except ImportError:
            messagebox.showwarning("Excel Not Available",
                                    "Install openpyxl:\n"
                                    "pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export: {e}")

    def refresh(self):
        try:
            self._calc_summary()
            self._calc_comparison()
        except Exception as e:
            app_log(f"[refresh] {e}")

