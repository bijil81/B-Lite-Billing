"""
Phase 4 Export Engine
All exports: transactions, GST summary, payment summary, daily closing,
profit summary, customer ledger, supplier ledger, improved Excel, improved PDF.
"""
from __future__ import annotations

import csv
import os
from collections import defaultdict
from datetime import datetime

from utils import DATA_DIR, F_REPORT, fmt_currency, safe_float, app_log, today_str, open_file_cross_platform
from date_helpers import display_to_iso_date, validate_display_date
from db_core.connection import connection_scope
from db_core.schema_manager import ensure_v5_schema
from repositories.reports_repo import ReportsRepository

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from expenses import get_expenses
except ImportError:
    def get_expenses():
        return []

try:
    from customers import get_customers
except ImportError:
    def get_customers():
        return {}

try:
    from inventory import get_stock
except ImportError:
    def get_stock():
        return []


_reports_repo = ReportsRepository()


# ───────────────────────── helpers ─────────────────────────

def _resolve_date_range(from_d="", to_d=""):
    """Convert DD-MM-YYYY display dates to YYYY-MM-DD ISO for DB queries."""
    iso_from = iso_to = None
    if from_d:
        if validate_display_date(from_d):
            iso_from = display_to_iso_date(from_d)
        else:
            iso_from = from_d  # already ISO-ish
    if to_d:
        if validate_display_date(to_d):
            iso_to = display_to_iso_date(to_d)
        else:
            iso_to = to_d
    return iso_from, iso_to


def _invoice_rows(from_d="", to_d="", search="") -> list[dict]:
    """Fetch report rows from v5 DB or fallback CSV."""
    from reports_data import read_report_rows
    return read_report_rows(from_d=from_d or "", to_d=to_d or "", search=search or "")


def _db_invoice_items(from_d="", to_d="") -> list[dict]:
    """Full transaction-level rows from DB (with per-line-item detail)."""
    ensure_v5_schema()
    query = """
        SELECT i.invoice_no,
               i.invoice_date,
               i.customer_name,
               i.customer_phone,
               i.net_total,
               i.discount_total,
               i.tax_total,
               ii.item_type,
               ii.item_name,
               ii.unit_price,
               ii.qty,
               ii.line_total,
               COALESCE((SELECT p.payment_method
                         FROM v5_payments p WHERE p.invoice_id = i.id
                         ORDER BY p.id LIMIT 1), '') AS payment_method,
               COALESCE(i.created_by, '') AS created_by
        FROM v5_invoices i
        JOIN v5_invoice_items ii ON ii.invoice_id = i.id
        WHERE COALESCE(i.is_deleted, 0) = 0
    """
    params: list = []
    if from_d:
        query += " AND i.invoice_date >= ?"
        params.append(from_d)
    if to_d:
        query += " AND i.invoice_date <= ?"
        params.append(to_d)
    query += " ORDER BY i.invoice_date DESC, i.id DESC, ii.id"

    with connection_scope() as conn:
        rows = conn.execute(query, tuple(params)).fetchall()
    return [dict(r) for r in rows]


def _require_openpyxl():
    if not HAS_OPENPYXL:
        raise RuntimeError("Excel export requires openpyxl, but it is not available in this build.")


# ─── Styling presets (Excel) ───

def _style_header(ws, col_count):
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1a1a2e")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for ci in range(1, col_count + 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def _style_data_row(ws, row, col_count):
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for ci in range(1, col_count + 1):
        cell = ws.cell(row=row, column=ci)
        cell.border = thin_border


def _auto_width(ws, col_count, max_w=50):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for c in col:
            val = str(c.value or "")
            # Account for currency/percent formatting
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 4, max_w)


# ───────────────────────── 1. Transaction-Level Export ─────────────────────────

def export_transactions_excel(from_d="", to_d="", search=""):
    """Export every line-item as a separate row in Excel."""
    _require_openpyxl()
    iso_from, iso_to = _resolve_date_range(from_d, to_d)
    items = _db_invoice_items(iso_from, iso_to)

    if not items:
        rows = _invoice_rows(iso_from or "", iso_to or "", search)
        if not rows:
            return _write_csv_fallback("transactions", ["Date","Invoice","Customer","Phone","Payment","Discount","Tax","Total"],
                                        [dict(d="Date",i=r.get("date",""),inv=r.get("invoice",""),nm=r.get("name",""),
                                              ph=r.get("phone",""),pm=r.get("payment",""),disc=r.get("discount",0),
                                              tax=0,tot=r.get("total",0)) for r in rows],
                                        lambda r: [r["i"],r["inv"],r["nm"],r["ph"],r["pm"],r["disc"],r["tax"],r["tot"]])

    out = os.path.join(DATA_DIR, f"transactions_{iso_from or 'all'}_to_{iso_to or 'all'}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Invoice", "Customer", "Phone", "Item Type", "Item Name",
               "Unit Price", "Qty", "Line Total", "Discount", "Tax", "Net Total", "Payment", "Cashier"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    for ri, it in enumerate(items, 2):
        vals = [
            str(it.get("invoice_date", "")),
            str(it.get("invoice_no", "")),
            str(it.get("customer_name", "")),
            str(it.get("customer_phone", "")),
            str(it.get("item_type", "")),
            str(it.get("item_name", "")),
            safe_float(it.get("unit_price", 0)),
            safe_float(it.get("qty", 0)),
            safe_float(it.get("line_total", 0)),
            safe_float(it.get("discount_total", 0)),
            safe_float(it.get("tax_total", 0)),
            safe_float(it.get("net_total", 0)),
            str(it.get("payment_method", "")),
            str(it.get("created_by", "")),
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v)
        _style_data_row(ws, ri, len(headers))

    _auto_width(ws, len(headers))

    # Freeze first row
    ws.freeze_panes = "A2"
    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"

    wb.save(out)
    return out


# ───────────────────────── 2. GST Summary Export ─────────────────────────

def export_gst_summary_excel(from_d="", to_d=""):
    """Export GST summary: taxable amount, CGST, SGST, total tax per invoice."""
    _require_openpyxl()
    iso_from, iso_to = _resolve_date_range(from_d, to_d)
    ensure_v5_schema()

    query = """
        SELECT i.invoice_no,
               i.invoice_date,
               i.customer_name,
               i.customer_phone,
               i.net_total,
               i.tax_total,
               i.discount_total
        FROM v5_invoices i
        WHERE COALESCE(i.is_deleted, 0) = 0
    """
    params: list = []
    if iso_from:
        query += " AND i.invoice_date >= ?"
        params.append(iso_from)
    if iso_to:
        query += " AND i.invoice_date <= ?"
        params.append(iso_to)
    query += " ORDER BY i.invoice_date"

    with connection_scope() as conn:
        rows = conn.execute(query, tuple(params)).fetchall()
    invoices = [dict(r) for r in rows]

    if not invoices:
        return _write_csv_fallback("gst_summary",
                                    ["Date","Invoice","Customer","Phone","Net","Tax","GST Rate%","CGST","SGST","Gross"],
                                    [],
                                    lambda r: [])

    out = os.path.join(DATA_DIR, f"gst_summary_{iso_from or 'all'}_to_{iso_to or 'all'}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GST Summary"

    headers = ["Date", "Invoice", "Customer", "Phone", "Net (Taxable)", "Tax Total",
               "Effective GST%", "CGST", "SGST", "Gross (Incl. Tax)"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    total_net = total_tax = total_gross = 0.0
    for ri, inv in enumerate(invoices, 2):
        net = safe_float(inv.get("net_total", 0))
        tax = safe_float(inv.get("tax_total", 0))
        gross = net + tax
        gst_pct = (tax / net * 100) if net > 0 else 0
        cgst = tax / 2
        sgst = tax / 2

        total_net += net
        total_tax += tax
        total_gross += gross

        vals = [
            str(inv.get("invoice_date", ""))[:10],
            str(inv.get("invoice_no", "")),
            str(inv.get("customer_name", "")),
            str(inv.get("customer_phone", "")),
            net, tax, f"{gst_pct:.2f}%", cgst, sgst, gross,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci in (5, 6, 8, 9, 10):
                cell.number_format = '#,##0.00'
        _style_data_row(ws, ri, len(headers))

    # Totals row
    tr = len(invoices) + 2
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=5, value=total_net)
    ws.cell(row=tr, column=6, value=total_tax)
    ws.cell(row=tr, column=8, value=total_tax / 2)
    ws.cell(row=tr, column=9, value=total_tax / 2)
    ws.cell(row=tr, column=10, value=total_gross)
    for ci in range(1, len(headers) + 1):
        ws.cell(row=tr, column=ci).font = Font(bold=True, size=11)
        ws.cell(row=tr, column=ci).fill = PatternFill("solid", fgColor="E8E8E8")
        if ci in (5, 6, 8, 9, 10):
            ws.cell(row=tr, column=ci).number_format = '#,##0.00'

    _auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    wb.save(out)
    return out


# ───────────────────────── 3. Payment Method Summary Export ─────────────────────────

def export_payment_summary_excel(from_d="", to_d=""):
    """Export payment method breakdown for a date range."""
    _require_openpyxl()
    iso_from, iso_to = _resolve_date_range(from_d, to_d)
    if iso_from and iso_to:
        breakdown = _reports_repo.payment_breakdown(iso_from, iso_to)
    else:
        ensure_v5_schema()
        query = """
            SELECT p.payment_method,
                   COUNT(DISTINCT p.invoice_id) AS count_invoices,
                   COALESCE(SUM(p.amount), 0) AS total_amount
            FROM v5_payments p
            JOIN v5_invoices i ON i.id = p.invoice_id
            WHERE COALESCE(i.is_deleted, 0) = 0
        """
        params: list = []
        if iso_from:
            query += " AND i.invoice_date >= ?"
            params.append(iso_from)
        if iso_to:
            query += " AND i.invoice_date <= ?"
            params.append(iso_to)
        query += " GROUP BY p.payment_method ORDER BY total_amount DESC"

        with connection_scope() as conn:
            rows = conn.execute(query, tuple(params)).fetchall()
        breakdown = [dict(r) for r in rows]

    if not breakdown:
        return _write_csv_fallback("payment_summary",
                                    ["Payment Method", "Invoice Count", "Total Amount"],
                                    [],
                                    lambda r: [])

    out = os.path.join(DATA_DIR, f"payment_summary_{iso_from or 'all'}_to_{iso_to or 'all'}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Summary"

    headers = ["Payment Method", "Invoice Count", "Total Amount", "% of Total"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    grand = sum(safe_float(r.get("total_amount", r.get("amount_total", 0))) for r in breakdown)
    total_count = 0
    for ri, row in enumerate(breakdown, 2):
        amt = safe_float(row.get("total_amount", row.get("amount_total", 0)))
        cnt = int(safe_float(row.get("count_invoices", 0)))
        pct = (amt / grand * 100) if grand > 0 else 0
        total_count += cnt
        for ci, v in enumerate([
            str(row["payment_method"]),
            cnt,
            amt,
            f"{pct:.1f}%",
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci == 3:
                cell.number_format = '#,##0.00'
        _style_data_row(ws, ri, len(headers))

    tr = len(breakdown) + 2
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=2, value=total_count)
    ws.cell(row=tr, column=3, value=grand)
    ws.cell(row=tr, column=4, value="100.0%")
    for ci in range(1, len(headers) + 1):
        ws.cell(row=tr, column=ci).font = Font(bold=True, size=11)
        ws.cell(row=tr, column=ci).fill = PatternFill("solid", fgColor="E8E8E8")
        if ci == 3:
            ws.cell(row=tr, column=ci).number_format = '#,##0.00'

    _auto_width(ws, len(headers))
    wb.save(out)
    return out


# ───────────────────────── 4. Daily Closing Export ─────────────────────────

def export_daily_closing_excel(dt=""):
    """Export daily closing report (bills + expenses + profit) to Excel."""
    _require_openpyxl()
    if not dt:
        from date_helpers import today_display_str
        dt = today_display_str()
    if validate_display_date(dt):
        dt_iso = display_to_iso_date(dt)
    else:
        dt_iso = dt

    # Bills for the day
    ensure_v5_schema()
    rows = []
    try:
        with connection_scope() as conn:
            bill_rows = conn.execute("""
                SELECT i.invoice_no, i.invoice_date, i.customer_name, i.customer_phone,
                       COALESCE((SELECT p.payment_method FROM v5_payments p
                                 WHERE p.invoice_id = i.id ORDER BY p.id LIMIT 1), '') AS payment,
                       i.net_total, i.discount_total
                FROM v5_invoices i
                WHERE substr(i.invoice_date, 1, 10) = ?
                  AND COALESCE(i.is_deleted, 0) = 0
                ORDER BY i.invoice_date
            """, (dt_iso,)).fetchall()
            rows = [dict(r) for r in bill_rows]
    except Exception as e:
        app_log(f"[export_daily_closing_excel:db] {e}")

    expenses = [e for e in get_expenses() if e.get("date", "")[:10] == dt_iso]

    total_rev = sum(r["net_total"] for r in rows)
    total_disc = sum(r["discount_total"] for r in rows)
    total_exp = sum(safe_float(e.get("amount", 0)) for e in expenses)
    net_profit = total_rev - total_exp

    payment_sum = defaultdict(float)
    for r in rows:
        payment_sum[r["payment"]] += r["net_total"]

    out = os.path.join(DATA_DIR, f"daily_closing_{dt_iso}.xlsx")
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Daily Summary"
    ws1.cell(row=1, column=1, value=f"Daily Closing Report - {dt}")
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws1.merge_cells("A1:B1")

    summary_data = [
        ("Total Bills", len(rows)),
        ("Gross Revenue", total_rev),
        ("Total Discount", total_disc),
        ("Total Expenses", total_exp),
        ("Net Profit", net_profit),
    ]
    r = 3
    for lbl, val in summary_data:
        ws1.cell(row=r, column=1, value=lbl)
        ws1.cell(row=r, column=1).font = Font(bold=True)
        c = ws1.cell(row=r, column=2, value=val)
        if isinstance(val, float):
            c.number_format = '#,##0.00'
        r += 1

    r += 1
    ws1.cell(row=r, column=1, value="Payment Breakdown")
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    for pm, amt in payment_sum.items():
        ws1.cell(row=r, column=1, value=f"- {pm}")
        c = ws1.cell(row=r, column=2, value=amt)
        c.number_format = '#,##0.00'
        r += 1

    # Sheet 2: Bill Details
    ws2 = wb.create_sheet("Bill Details")
    headers = ["Invoice", "Time", "Customer", "Phone", "Payment", "Discount", "Total"]
    for ci, h in enumerate(headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header(ws2, len(headers))

    for ri, b in enumerate(rows, 2):
        time_str = str(b.get("invoice_date", ""))[11:16] if len(str(b.get("invoice_date", ""))) >= 16 else ""
        vals = [
            str(b.get("invoice_no", "")),
            time_str,
            str(b.get("customer_name", "")),
            str(b.get("customer_phone", "")),
            str(b.get("payment", "")),
            safe_float(b.get("discount_total", 0)),
            safe_float(b.get("net_total", 0)),
        ]
        for ci, v in enumerate(vals, 1):
            ws2.cell(row=ri, column=ci, value=v)
            if ci in (6, 7):
                ws2.cell(row=ri, column=ci).number_format = '#,##0.00'
        _style_data_row(ws2, ri, len(headers))

    # Sheet 3: Expenses
    if expenses:
        ws3 = wb.create_sheet("Expenses")
        for ci, h in enumerate(["Category", "Description", "Amount", "Date"], 1):
            ws3.cell(row=1, column=ci, value=h)
        _style_header(ws3, 4)
        for ri, e in enumerate(expenses, 2):
            for ci, v in enumerate([
                e.get("category", ""),
                e.get("description", ""),
                safe_float(e.get("amount", 0)),
                e.get("date", ""),
            ], 1):
                ws3.cell(row=ri, column=ci, value=v)
                if ci == 3:
                    ws3.cell(row=ri, column=ci).number_format = '#,##0.00'
            _style_data_row(ws3, ri, 4)

    _auto_width(ws2, len(headers))
    wb.save(out)
    return out


# ───────────────────────── 5. Profit Summary Export ─────────────────────────

def export_profit_summary_excel(from_d="", to_d=""):
    """Export service-level profit analysis (revenue by service)."""
    _require_openpyxl()
    iso_from, iso_to = _resolve_date_range(from_d, to_d)

    svc_rev = defaultdict(float)
    svc_count = defaultdict(int)
    all_total = 0.0

    ensure_v5_schema()
    try:
        with connection_scope() as conn:
            rows = conn.execute("""
                SELECT ii.item_name, ii.unit_price, ii.qty
                FROM v5_invoice_items ii
                JOIN v5_invoices i ON i.id = ii.invoice_id
                WHERE ii.item_type = 'services' AND COALESCE(i.is_deleted, 0) = 0
            """ + ("" if not iso_from else """ AND i.invoice_date >= ?""" )
              + ("" if not iso_to else """ AND i.invoice_date <= ?"""),
                tuple(filter(None, [iso_from, iso_to]))
            ).fetchall()
            for r in rows:
                amt = safe_float(r["unit_price"]) * safe_float(r["qty"])
                svc_rev[r["item_name"]] += amt
                svc_count[r["item_name"]] += int(safe_float(r["qty"]))
                all_total += amt
    except Exception as e:
        app_log(f"[export_profit_summary:db] {e}")

    if not svc_rev:
        return _write_csv_fallback("profit_summary",
                                    ["Service", "Times Sold", "Total Revenue", "% of Revenue"],
                                    [],
                                    lambda r: [])

    out = os.path.join(DATA_DIR, f"profit_summary_{iso_from or 'all'}_to_{iso_to or 'all'}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit Analysis"

    headers = ["Service", "Times Sold", "Total Revenue", "% of Revenue"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    sorted_svcs = sorted(svc_rev.items(), key=lambda x: x[1], reverse=True)
    for ri, (svc, rev) in enumerate(sorted_svcs, 2):
        pct = (rev / all_total * 100) if all_total > 0 else 0
        for ci, v in enumerate([svc, svc_count[svc], rev, f"{pct:.1f}%"], 1):
            ws.cell(row=ri, column=ci, value=v)
            if ci == 3:
                ws.cell(row=ri, column=ci).number_format = '#,##0.00'
        _style_data_row(ws, ri, len(headers))

    tr = len(sorted_svcs) + 2
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=2, value=sum(svc_count.values()))
    ws.cell(row=tr, column=3, value=all_total)
    ws.cell(row=tr, column=4, value="100.0%")
    for ci in range(1, len(headers) + 1):
        ws.cell(row=tr, column=ci).font = Font(bold=True, size=11)
        ws.cell(row=tr, column=ci).fill = PatternFill("solid", fgColor="E8E8E8")
        if ci == 3:
            ws.cell(row=tr, column=ci).number_format = '#,##0.00'

    _auto_width(ws, len(headers))
    wb.save(out)
    return out


# ───────────────────────── 6. Customer Ledger Export ─────────────────────────

def export_customer_ledger_excel(customer_phone="", from_d="", to_d=""):
    """Export all transactions for a specific customer (or all customers)."""
    _require_openpyxl()
    iso_from, iso_to = _resolve_date_range(from_d, to_d)

    ensure_v5_schema()
    query = """
        SELECT i.invoice_no, i.invoice_date, i.customer_name, i.customer_phone,
               i.net_total, i.discount_total, i.tax_total,
               COALESCE((SELECT p.payment_method FROM v5_payments p
                         WHERE p.invoice_id = i.id ORDER BY p.id LIMIT 1), '') AS payment
        FROM v5_invoices i
        WHERE COALESCE(i.is_deleted, 0) = 0
    """
    params: list = []
    if customer_phone:
        query += " AND i.customer_phone = ?"
        params.append(customer_phone)
    if iso_from:
        query += " AND i.invoice_date >= ?"
        params.append(iso_from)
    if iso_to:
        query += " AND i.invoice_date <= ?"
        params.append(iso_to)
    query += " ORDER BY i.invoice_date"

    with connection_scope() as conn:
        rows = conn.execute(query, tuple(params)).fetchall()
    invoices = [dict(r) for r in rows]

    if not invoices:
        return _write_csv_fallback("customer_ledger",
                                    ["Date","Invoice","Customer","Phone","Payment","Net","Discount","Tax","Total"],
                                    [],
                                    lambda r: [])

    label = customer_phone or "all_customers"
    label = label.replace(" ", "_")
    out = os.path.join(DATA_DIR, f"customer_ledger_{label}_{iso_from or 'all'}_to_{iso_to or 'all'}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"

    headers = ["Date", "Invoice", "Customer", "Phone", "Payment", "Net", "Discount", "Tax", "Total"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    running = 0.0
    for ri, inv in enumerate(invoices, 2):
        net = safe_float(inv.get("net_total", 0))
        disc = safe_float(inv.get("discount_total", 0))
        tax = safe_float(inv.get("tax_total", 0))
        total = net + tax
        running += total
        date_str = str(inv.get("invoice_date", ""))[:10]
        vals = [date_str, str(inv.get("invoice_no", "")),
                str(inv.get("customer_name", "")), str(inv.get("customer_phone", "")),
                str(inv.get("payment", "")), net, disc, tax, total]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci in (6, 7, 8, 9):
                cell.number_format = '#,##0.00'
        _style_data_row(ws, ri, len(headers))

        c = ws.cell(row=ri, column=10, value=round(running, 2))
        c.number_format = '#,##0.00'

    # Running balance column
    ws.cell(row=1, column=10, value="Running Balance")
    ws.cell(row=1, column=10).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=10).fill = PatternFill("solid", fgColor="1a1a2e")

    tr = len(invoices) + 2
    grand = sum(safe_float(i.get("net_total", 0)) + safe_float(i.get("tax_total", 0)) for i in invoices)
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=9, value=grand)
    ws.cell(row=tr, column=9).number_format = '#,##0.00'
    ws.cell(row=tr, column=9).font = Font(bold=True, size=11)
    ws.cell(row=tr, column=9).fill = PatternFill("solid", fgColor="E8E8E8")

    _auto_width(ws, len(headers) + 1)
    ws.freeze_panes = "A2"
    wb.save(out)
    return out


# ───────────────────────── 7. Supplier Ledger Export ─────────────────────────

def export_supplier_ledger_excel(from_d="", to_d=""):
    """Export supplier/purchase ledger from inventory purchases."""
    _require_openpyxl()
    iso_from, iso_to = _resolve_date_range(from_d, to_d)

    ensure_v5_schema()
    try:
        with connection_scope() as conn:
            tables = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
            table_names = [r["name"] for r in tables]
    except Exception:
        table_names = []

    if "v5_inventory_purchases" not in table_names:
        return None

    purchases = []
    try:
        with connection_scope() as conn:
            rows = conn.execute("""
                SELECT p.id, p.purchase_date, p.supplier_name, p.supplier_phone,
                       p.item_name, p.qty, p.unit_cost, p.total_cost,
                       COALESCE(p.payment_method, 'Credit') AS payment_method,
                       COALESCE(p.status, 'received') AS status
                FROM v5_inventory_purchases p
                WHERE 1=1
            """ + ("" if not iso_from else " AND p.purchase_date >= ?")
              + ("" if not iso_to else " AND p.purchase_date <= ?"),
                tuple(filter(None, [iso_from, iso_to]))
            ).fetchall()
            purchases = [dict(r) for r in rows]
    except Exception as e:
        app_log(f"[export_supplier_ledger] {e}")

    if not purchases:
        return _write_csv_fallback("supplier_ledger",
                                    ["Date","Supplier","Item","Qty","Unit Cost","Total","Payment","Status"],
                                    [],
                                    lambda r: [])

    out = os.path.join(DATA_DIR, f"supplier_ledger_{iso_from or 'all'}_to_{iso_to or 'all'}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Ledger"

    headers = ["Date", "Supplier", "Item", "Qty", "Unit Cost", "Total Cost", "Payment", "Status"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    total_cost = 0.0
    for ri, p in enumerate(purchases, 2):
        vals = [
            str(p.get("purchase_date", ""))[:10],
            str(p.get("supplier_name", "Unknown")),
            str(p.get("item_name", "")),
            int(safe_float(p.get("qty", 0))),
            safe_float(p.get("unit_cost", 0)),
            safe_float(p.get("total_cost", 0)),
            str(p.get("payment_method", "")),
            str(p.get("status", "")),
        ]
        total_cost += safe_float(p.get("total_cost", 0))
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci in (5, 6):
                cell.number_format = '#,##0.00'
        _style_data_row(ws, ri, len(headers))

    tr = len(purchases) + 2
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=6, value=total_cost)
    ws.cell(row=tr, column=6).number_format = '#,##0.00'
    ws.cell(row=tr, column=6).font = Font(bold=True, size=11)
    ws.cell(row=tr, column=6).fill = PatternFill("solid", fgColor="E8E8E8")

    _auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    wb.save(out)
    return out


# ───────────────────────── 8. Improved Sales Excel Export ─────────────────────────

def export_sales_report_excel(from_d="", to_d="", search=""):
    """Improved, better-formatted sales report Excel export."""
    _require_openpyxl()
    iso_from, iso_to = _resolve_date_range(from_d, to_d)
    rows = _invoice_rows(iso_from or "", iso_to or "", search)

    if not rows:
        return _write_csv_fallback("sales_report",
                                    ["Date","Invoice","Customer","Phone","Payment","Discount","Tax","Total","Items"],
                                    [],
                                    lambda r: [])

    out = os.path.join(DATA_DIR, f"sales_report_{iso_from or 'all'}_to_{iso_to or 'all'}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ["Date", "Invoice", "Customer", "Phone", "Payment", "Discount", "Total", "Items"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, len(headers))

    total_disc = total_rev = 0.0
    for ri, row in enumerate(rows, 2):
        items_readable = _format_items_for_export(row.get("items_raw", ""))
        disc = safe_float(row.get("discount", 0))
        total = safe_float(row.get("total", 0))
        total_disc += disc
        total_rev += total

        for ci, v in enumerate([
            str(row.get("date", "")),
            str(row.get("invoice", "")),
            str(row.get("name", "")),
            str(row.get("phone", "")),
            str(row.get("payment", "")),
            disc,
            total,
            items_readable,
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci in (6, 7):
                cell.number_format = '#,##0.00'
        _style_data_row(ws, ri, len(headers))

    tr = len(rows) + 2
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=6, value=total_disc)
    ws.cell(row=tr, column=7, value=total_rev)
    for ci in (6, 7):
        ws.cell(row=tr, column=ci).number_format = '#,##0.00'
        ws.cell(row=tr, column=ci).font = Font(bold=True, size=11)
        ws.cell(row=tr, column=ci).fill = PatternFill("solid", fgColor="E8E8E8")

    ws.column_dimensions["H"].width = 60  # Items column wider
    _auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    wb.save(out)
    return out


def _format_items_for_export(items_raw: str) -> str:
    """Format items_raw into human-readable string."""
    if not items_raw:
        return ""
    parts_list = []
    for seg in items_raw.split("|"):
        seg = seg.strip()
        if not seg:
            continue
        # New format: mode~name~price~qty
        p = seg.split("~")
        if len(p) >= 4:
            try:
                name = p[1].strip()
                qty = int(float(p[3].strip()))
                parts_list.append(f"{name} x{qty}")
            except Exception:
                parts_list.append(seg)
        else:
            parts_list.append(seg)
    return " | ".join(parts_list)


# ───────────────────────── 9. Improved PDF Export ─────────────────────────

def export_sales_report_pdf(from_d="", to_d="", search=""):
    """Generate a professional PDF sales report."""
    if not HAS_REPORTLAB:
        return None

    iso_from, iso_to = _resolve_date_range(from_d, to_d)
    rows = _invoice_rows(iso_from or "", iso_to or "", search)
    if not rows:
        return None

    date_label = f"{iso_from or 'Start'} to {iso_to or 'End'}"
    out_path = os.path.join(DATA_DIR, f"sales_report_{iso_from or 'all'}_to_{iso_to or 'all'}.pdf")

    try:
        from branding import get_company_name, get_invoice_branding
        brand = get_invoice_branding()
    except Exception:
        brand = {"header": "Sales Report"}
        try:
            from branding import get_company_name
        except Exception:
            pass

    c = canvas.Canvas(out_path, pagesize=A4)
    W, H = A4
    y = H - 40

    # Header
    c.setFillColorRGB(0.1, 0.1, 0.18)
    c.rect(0, H - 65, W, 65, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(W / 2, H - 32, brand.get("header", "Sales Report"))
    c.setFont("Helvetica", 10)
    c.drawCentredString(W / 2, H - 50, f"Sales Report  |  {date_label}")

    y = H - 80

    def _line(text, bold=False, size=10, color=(0, 0, 0)):
        nonlocal y
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.setFillColorRGB(*color)
        c.drawString(40, y, text)
        y -= size + 4

    def _divider():
        nonlocal y
        c.setStrokeColorRGB(0.85, 0.85, 0.85)
        c.line(40, y, W - 40, y)
        y -= 8

    # Table header
    _line("INVOICE DETAILS", bold=True, size=12, color=(0.1, 0.1, 0.5))
    _divider()

    hdr_font_size = 8
    col_x = [40, 100, 200, 330, 420, 500]
    hdrs = ["Date", "Invoice", "Customer", "Phone", "Payment", "Total"]
    c.setFont("Helvetica-Bold", hdr_font_size)
    c.setFillColorRGB(0.3, 0.3, 0.3)
    for hx, ht in zip(col_x, hdrs):
        c.drawString(hx, y, ht)
    y -= 12
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.5)
    c.line(40, y, W - 40, y)
    y -= 6

    total_rev = 0.0
    for r in rows:
        if y < 50:
            c.showPage()
            y = H - 40
        vals = [
            str(r.get("date", ""))[:10],
            str(r.get("invoice", "")),
            str(r.get("name", ""))[:18],
            str(r.get("phone", "")),
            str(r.get("payment", "")),
            fmt_currency(safe_float(r.get("total", 0))),
        ]
        c.setFont("Helvetica", hdr_font_size)
        c.setFillColorRGB(0, 0, 0)
        for hx, vt in zip(col_x, vals):
            c.drawString(hx, y, vt)
        y -= 12
        total_rev += safe_float(r.get("total", 0))

    y -= 10
    _divider()
    _line(f"TOTAL REVENUE: {fmt_currency(total_rev)}", bold=True, size=11, color=(0, 0.5, 0.2))
    _line(f"TOTAL INVOICES: {len(rows)}", bold=True, size=10, color=(0.1, 0.1, 0.5))

    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    try:
        cn = get_company_name()
    except Exception:
        cn = ""
    c.drawCentredString(W / 2, 25,
                         f"Generated: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}  |  {cn}")
    c.save()
    return out_path


# ───────────────────────── CSV Fallback ─────────────────────────

def _write_csv_fallback(name, headers, data_rows, row_extractor):
    """Write CSV as fallback when no data found."""
    out = os.path.join(DATA_DIR, f"{name}_{today_str()}.csv")
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow(row_extractor(row))
    return out
