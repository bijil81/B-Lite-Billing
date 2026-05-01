"""Bulk product import parsing, validation, and preview helpers.

G6A is intentionally pure: it does not write inventory or catalog rows.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
import json
from pathlib import Path
from typing import Any, Iterable, Mapping

from .product_form import build_inventory_product_form_payload


CANONICAL_FIELDS = (
    "name",
    "category",
    "brand",
    "base_product",
    "pack_size",
    "bill_label",
    "barcode",
    "sku",
    "unit",
    "sale_unit",
    "base_unit",
    "qty",
    "min_stock",
    "cost",
    "sale_price",
    "mrp",
    "gst_rate",
    "hsn_sac",
    "price_includes_tax",
    "allow_decimal_qty",
    "is_weighed",
)


COLUMN_ALIASES = {
    "name": {"name", "item", "item name", "product", "product name", "product_name"},
    "category": {"category", "cat", "product category"},
    "brand": {"brand", "brand name"},
    "base_product": {"base product", "base_product", "base name", "base_name"},
    "pack_size": {"pack", "pack size", "pack_size", "variant", "size"},
    "bill_label": {"bill label", "bill_label", "billing name", "display name"},
    "barcode": {"barcode", "bar code", "ean", "upc"},
    "sku": {"sku", "item code", "item_code", "code"},
    "unit": {"unit", "unit type", "unit_type", "measurement", "unit (measurement)"},
    "sale_unit": {"sale unit", "sale_unit", "price basis", "price_basis"},
    "base_unit": {"base unit", "base_unit"},
    "qty": {"qty", "quantity", "stock", "opening stock", "opening_stock", "stock qty"},
    "min_stock": {"min stock", "min_stock", "reorder", "reorder level", "reorder_level"},
    "cost": {"cost", "cost price", "cost_price", "purchase price", "purchase_price"},
    "sale_price": {"sale price", "sale_price", "selling price", "selling_price", "price", "rate"},
    "mrp": {"mrp", "maximum retail price"},
    "gst_rate": {"gst", "gst rate", "gst_rate", "gst %", "gst rate %", "tax", "tax rate"},
    "hsn_sac": {"hsn", "hsn/sac", "hsn_sac", "sac"},
    "price_includes_tax": {"tax inclusive", "tax_inclusive", "price includes tax", "price_includes_tax"},
    "allow_decimal_qty": {"allow decimal", "allow_decimal_qty", "decimal qty"},
    "is_weighed": {"weighed", "is weighed", "is_weighed", "loose"},
}


REQUIRED_FIELDS = ("name", "category", "sale_price")


@dataclass(frozen=True)
class ImportIssue:
    row_number: int
    field: str
    message: str
    severity: str = "error"


@dataclass(frozen=True)
class ImportPreviewRow:
    row_number: int
    action: str
    raw: dict[str, Any]
    mapped: dict[str, Any]
    inventory_item: dict[str, Any] = field(default_factory=dict)
    catalog_payload: dict[str, Any] = field(default_factory=dict)
    match_key: str = ""
    errors: tuple[ImportIssue, ...] = ()
    warnings: tuple[ImportIssue, ...] = ()


@dataclass(frozen=True)
class ImportPreview:
    rows: tuple[ImportPreviewRow, ...]
    errors: tuple[ImportIssue, ...] = ()

    @property
    def created_count(self) -> int:
        return sum(1 for row in self.rows if row.action == "create")

    @property
    def updated_count(self) -> int:
        return sum(1 for row in self.rows if row.action == "update")

    @property
    def skipped_count(self) -> int:
        return sum(1 for row in self.rows if row.action == "skip")

    @property
    def error_count(self) -> int:
        return len(self.errors) + sum(len(row.errors) for row in self.rows)

    @property
    def warning_count(self) -> int:
        return sum(len(row.warnings) for row in self.rows)

    @property
    def ok_to_import(self) -> bool:
        return self.error_count == 0 and bool(self.rows)


def _normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().replace("_", " ").lower().split())


def default_column_mapping(headers: Iterable[object]) -> dict[str, str]:
    result: dict[str, str] = {}
    alias_to_field = {
        alias: field_name
        for field_name, aliases in COLUMN_ALIASES.items()
        for alias in aliases
    }
    for header in headers:
        normalized = _normalize_header(header)
        field_name = alias_to_field.get(normalized)
        if field_name and field_name not in result:
            result[field_name] = str(header)
    return result


def _clean_value(value: object) -> str:
    return str(value if value is not None else "").strip()


def map_import_row(row: Mapping[str, Any], mapping: Mapping[str, str] | None = None) -> dict[str, Any]:
    source = dict(row)
    active_mapping = dict(mapping or default_column_mapping(source.keys()))
    mapped: dict[str, Any] = {}
    for field_name in CANONICAL_FIELDS:
        source_column = active_mapping.get(field_name)
        if not source_column:
            continue
        mapped[field_name] = _clean_value(source.get(source_column, ""))
    return mapped


def parse_csv_text(text: str) -> list[dict[str, Any]]:
    reader = csv.DictReader((text or "").splitlines())
    if not reader.fieldnames:
        return []
    return [dict(row) for row in reader]


def parse_json_text(text: str) -> list[dict[str, Any]]:
    if not str(text or "").strip():
        return []
    data = json.loads(text)
    if isinstance(data, dict):
        data = data.get("products") or data.get("items") or data.get("rows") or []
    if not isinstance(data, list):
        raise ValueError("JSON import must contain a list of product rows")
    return [dict(row) for row in data if isinstance(row, Mapping)]


def parse_xlsx_file(path: str | Path, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for row in rows[1:]:
        values = {header: value for header, value in zip(headers, row) if header}
        if any(_clean_value(value) for value in values.values()):
            result.append(values)
    return result


def parse_import_file(path: str | Path) -> list[dict[str, Any]]:
    target = Path(path)
    suffix = target.suffix.lower()
    if suffix == ".csv":
        return parse_csv_text(target.read_text(encoding="utf-8-sig"))
    if suffix == ".json":
        return parse_json_text(target.read_text(encoding="utf-8"))
    if suffix == ".xlsx":
        return parse_xlsx_file(target)
    raise ValueError(f"Unsupported import file type: {suffix or 'unknown'}")


def _existing_indexes(existing_items: Mapping[str, Mapping[str, Any]] | Iterable[Mapping[str, Any]] | None) -> dict[str, dict[str, str]]:
    indexes = {"barcode": {}, "sku": {}, "name": {}}
    if not existing_items:
        return indexes
    if isinstance(existing_items, Mapping):
        iterable = [
            {"name": name, **dict(item or {})}
            for name, item in existing_items.items()
        ]
    else:
        iterable = [dict(item or {}) for item in existing_items]
    for item in iterable:
        name = _clean_value(item.get("name") or item.get("legacy_name") or item.get("bill_label"))
        barcode = _clean_value(item.get("barcode"))
        sku = _clean_value(item.get("sku")).upper()
        if barcode:
            indexes["barcode"][barcode] = name
        if sku:
            indexes["sku"][sku] = name
        if name:
            indexes["name"][name.lower()] = name
    return indexes


def _match_existing(mapped: Mapping[str, Any], indexes: Mapping[str, Mapping[str, str]]) -> tuple[str, str]:
    barcode = _clean_value(mapped.get("barcode"))
    sku = _clean_value(mapped.get("sku")).upper()
    name = _clean_value(mapped.get("name")).lower()
    if barcode and barcode in indexes["barcode"]:
        return "barcode", indexes["barcode"][barcode]
    if sku and sku in indexes["sku"]:
        return "sku", indexes["sku"][sku]
    if name and name in indexes["name"]:
        return "name", indexes["name"][name]
    return "", ""


def build_import_preview(
    raw_rows: Iterable[Mapping[str, Any]],
    *,
    column_mapping: Mapping[str, str] | None = None,
    existing_items: Mapping[str, Mapping[str, Any]] | Iterable[Mapping[str, Any]] | None = None,
    duplicate_policy: str = "skip",
    below_cost_policy: str = "warn",
) -> ImportPreview:
    rows = list(raw_rows or [])
    if not rows:
        return ImportPreview(rows=(), errors=(ImportIssue(0, "file", "Import file is empty"),))

    all_headers: list[str] = []
    seen_headers: set[str] = set()
    for row in rows:
        for header in row.keys():
            text = str(header)
            if text not in seen_headers:
                seen_headers.add(text)
                all_headers.append(text)
    mapping = dict(column_mapping or default_column_mapping(all_headers))
    preview_rows: list[ImportPreviewRow] = []
    existing = _existing_indexes(existing_items)
    seen_barcode: dict[str, int] = {}
    seen_sku: dict[str, int] = {}
    seen_name: dict[str, int] = {}

    for index, raw in enumerate(rows, start=2):
        mapped = map_import_row(raw, mapping)
        errors: list[ImportIssue] = []
        warnings: list[ImportIssue] = []
        for field_name in REQUIRED_FIELDS:
            if not _clean_value(mapped.get(field_name)):
                errors.append(ImportIssue(index, field_name, f"{field_name} is required"))

        barcode = _clean_value(mapped.get("barcode"))
        sku = _clean_value(mapped.get("sku")).upper()
        name_key = _clean_value(mapped.get("name")).lower()
        for field_name, value, seen in (
            ("barcode", barcode, seen_barcode),
            ("sku", sku, seen_sku),
            ("name", name_key, seen_name),
        ):
            if not value:
                continue
            if value in seen:
                errors.append(
                    ImportIssue(
                        index,
                        field_name,
                        f"Duplicate {field_name} also appears on row {seen[value]}",
                    )
                )
            else:
                seen[value] = index

        form_payload = None
        if not errors:
            form_payload = build_inventory_product_form_payload(mapped)
            for issue in form_payload.validation.errors:
                errors.append(ImportIssue(index, issue.field, issue.message))
            for issue in form_payload.validation.warnings:
                import_issue = ImportIssue(index, issue.field, issue.message, "warning")
                if below_cost_policy == "error":
                    errors.append(ImportIssue(index, issue.field, issue.message))
                elif below_cost_policy != "allow":
                    warnings.append(import_issue)

        match_type, matched_name = _match_existing(mapped, existing)
        match_key = f"{match_type}:{matched_name}" if match_type else ""
        action = "error" if errors else "create"
        if not errors and matched_name:
            if duplicate_policy == "update":
                action = "update"
            elif duplicate_policy == "error":
                errors.append(ImportIssue(index, match_type, f"Existing item matched: {matched_name}"))
                action = "error"
            else:
                warnings.append(ImportIssue(index, match_type, f"Existing item skipped: {matched_name}", "warning"))
                action = "skip"

        preview_rows.append(
            ImportPreviewRow(
                row_number=index,
                action=action,
                raw=dict(raw),
                mapped=mapped,
                inventory_item=form_payload.inventory_item if form_payload and not errors else {},
                catalog_payload=form_payload.catalog_payload if form_payload and not errors else {},
                match_key=match_key,
                errors=tuple(errors),
                warnings=tuple(warnings),
            )
        )

    return ImportPreview(rows=tuple(preview_rows))
