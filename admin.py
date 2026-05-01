"""
admin.py  —  BOBY'S Salon : Admin panel (services, products, users)
FIXES:
  - Bug 12: added current_user parameter to __init__
  - Bug 11: _close() with proper grab_release + protocol handler
  - Bug 10: smart search closure pattern is correct (kept as-is)
  - Fix R9a: app_log imported — errors logged to app_debug.log
  - Fix R9b: _load_db() try/except — crash prevention
  - Fix R9c: _save_db() try/except + app_log — save failure logged
  - Fix R9d: __init__ try/except — admin panel open failure logged
  - Fix R9e: _build() try/except — UI build failure logged
  - Fix R9f: add_cat() try/except — error shown to user
  - Fix R9g: del_cat() try/except — error shown to user
  - Fix R9h: load_items() try/except — crash prevention
  - Fix R9i: save_item() try/except — save failure logged + user notified
  - Fix R9j: delete_item() try/except — delete failure logged
  - Fix R9k: _build_users_tab() try/except — crash prevention
"""
import json
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from utils import (C, load_json, save_json, F_SERVICES,
                   hash_pw, center_window, popup_window, app_log)
from ui_theme import ModernButton
from ui_responsive import fit_toplevel, make_scrollable
from auth import UserManagerWindow
from icon_system import get_action_icon
from ui_responsive import get_responsive_metrics, scaled_value
from src.blite_v6.app.window_lifecycle import hide_while_building, reveal_when_ready
from adapters.product_catalog_adapter import (
    create_product_with_variants_v5,
    refresh_product_catalog_cache,
    search_product_variants_v5,
    use_v5_product_variants_db,
)
from help_system import show_context_help
from branding import get_app_name
from services_v5.inventory_service import InventoryService
from validators.product_validator import build_pack_label


def _load_db():
    """Fix R9b: try/except — returns empty dicts on failure."""
    try:
        data = load_json(F_SERVICES, {})
        if "Services" in data or "Products" in data:
            return data.get("Services", {}), data.get("Products", {})
        return data, {}
    except Exception as e:
        app_log(f"[_load_db] {e}")
        return {}, {}


def _save_db(svc: dict, prd: dict):
    """Fix R9c: try/except + app_log on save failure."""
    try:
        return save_json(F_SERVICES, {"Services": svc, "Products": prd})
    except Exception as e:
        app_log(f"[_save_db] {e}")
        return False


def _product_display_name(name: str, brand: str = "", pack_size: str = "", unit: str = "") -> str:
    parts = [str(brand or "").strip(), str(name or "").strip()]
    if str(pack_size or "").strip() or str(unit or "").strip():
        parts.append(build_pack_label(pack_size, unit))
    return " ".join(p for p in parts if p).strip()


def _normalize_import_rows(payload) -> list[dict]:
    rows = []
    if isinstance(payload, list):
        rows = payload
    elif isinstance(payload, dict):
        if isinstance(payload.get("items"), list):
            rows = payload["items"]
        elif isinstance(payload.get("Products"), dict):
            for category, items in payload["Products"].items():
                if isinstance(items, dict):
                    for item_name, price in items.items():
                        rows.append({
                            "Item Name": item_name,
                            "Category": category,
                            "Price": price,
                            "Stock": 0,
                            "Unit": "pcs",
                        })
    return [row for row in rows if isinstance(row, dict)]


def _show_import_summary(title: str, source_name: str, summary: dict):
    lines = [
        f"Source: {source_name}",
        f"Rows read: {summary.get('rows_read', 0)}",
        f"Imported: {summary.get('imported', 0)}",
        f"New items: {summary.get('created', 0)}",
        f"Updated items: {summary.get('updated', 0)}",
        f"Skipped rows: {summary.get('skipped', 0)}",
        f"Warnings / errors: {summary.get('errors', 0)}",
    ]
    error_lines = summary.get("error_lines", [])
    if error_lines:
        lines.append("")
        lines.append("Top issues:")
        lines.extend(error_lines[:5])
    messagebox.showinfo(title, "\n".join(lines))


def _inventory_sync_payload(
    category: str,
    item_name: str,
    price: float,
    brand: str = "",
    pack_size: str = "",
    unit: str = "pcs",
    stock_qty: int = 0,
) -> dict:
    display_name = _product_display_name(item_name, brand, pack_size, unit) or item_name
    return {
        "legacy_name": display_name,
        "category": category,
        "brand": brand,
        "unit": unit or "pcs",
        "current_qty": stock_qty,
        "min_qty": 0,
        "cost_price": float(price or 0.0),
        "sale_price": float(price or 0.0),
        "active": True,
        "bill_label": display_name,
        "base_product": item_name,
        "pack_size": pack_size,
    }


class AdminPanel:
    """Opens the admin panel as a Toplevel window."""

    def __init__(self, parent, on_close=None, current_user=None):
        # H11 FIX: Reject missing current_user instead of silently defaulting
        # to owner. Defaulting to owner was a privilege escalation risk -- if
        # any caller passed None (e.g., due to a bug), the user would get
        # full admin access without authentication.
        if not current_user or not isinstance(current_user, dict):
            raise ValueError(
                "AdminPanel requires a valid current_user dict. "
                "This is a programming error -- check the caller."
            )
        if not current_user.get("role") or not current_user.get("username"):
            raise ValueError(
                "AdminPanel current_user must include 'role' and 'username'. "
                f"Got: {current_user!r}"
            )
        # Bug 12 fix: accept and store current_user
        self.parent       = parent
        self.on_close     = on_close
        self._responsive  = get_responsive_metrics(parent.winfo_toplevel())
        self.current_user = current_user

        self.svc_data, self.prd_data = _load_db()

        self.win = tk.Toplevel(parent)
        hide_while_building(self.win)
        self.win.title(f"Admin Panel - {get_app_name()}")
        popup_window(self.win, 1100, 700)
        self.win.configure(bg=C["bg"])
        self.win.grab_set()
        # Bug 11 fix: always use _close for WM_DELETE_WINDOW
        self.win.protocol("WM_DELETE_WINDOW", self._close)

        try:
            self._build()
        except Exception as e:
            app_log(f"[AdminPanel.__init__] build failed: {e}")
        finally:
            reveal_when_ready(self.win)

    def _close(self):
        """Safe close: release grab, run callback, destroy window."""
        try:
            self.win.grab_release()
        except Exception:
            pass
        try:
            if callable(self.on_close):
                self.on_close()
        except Exception:
            pass
        try:
            self.win.destroy()
        except Exception:
            pass

    def _build(self):
        top = tk.Frame(self.win, bg=C["bg"])
        top.pack(fill=tk.X, padx=10, pady=(10, 0))
        ModernButton(
            top,
            text="Help",
            command=lambda: show_context_help(self.win, "admin_panel"),
            color=C["blue"],
            hover_color="#154360",
            width=scaled_value(100, 92, 80),
            height=scaled_value(30, 30, 26),
            radius=8,
            font=("Arial", scaled_value(10, 10, 9), "bold"),
        ).pack(side=tk.RIGHT)

        nb = ttk.Notebook(self.win)
        nb.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        t_svc = tk.Frame(nb, bg=C["bg"])
        t_prd = tk.Frame(nb, bg=C["bg"])
        t_usr = tk.Frame(nb, bg=C["bg"])
        t_bu  = tk.Frame(nb, bg=C["bg"])

        nb.add(t_svc, text="Services")
        nb.add(t_prd, text="Products")
        nb.add(t_usr, text="Users")
        nb.add(t_bu, text="Backup & Logs")

        self._build_tab(t_svc, self.svc_data, "Services")
        self._build_tab(t_prd, self.prd_data, "Products")
        self._build_users_tab(t_usr)
        self._build_backup_tab(t_bu)

    # ─────────── Services / Products tab ────────────
    def _build_tab(self, parent, data_ref: dict, label: str):
        content, _, _ = make_scrollable(parent, bg=C["bg"], padx=0, pady=0)

        top = tk.Frame(content, bg=C["bg"], pady=8)
        top.pack(fill=tk.X, padx=10)
        for col in range(4):
            top.grid_columnconfigure(col, weight=1 if col in (1, 3) else 0)

        tk.Label(top, text="Category:", font=("Arial", 12, "bold"),
                 bg=C["bg"], fg=C["text"]).grid(
                     row=0, column=0, padx=(0, 8), pady=4, sticky="w")
        cat_var = tk.StringVar()
        cat_cb  = ttk.Combobox(top, textvariable=cat_var,
                                state="readonly", width=22)
        cat_cb.grid(row=0, column=1, padx=(0, 12), pady=4, sticky="ew")
        cat_cb["values"] = ["All"]

        tk.Label(top, text="New Category:", font=("Arial", 12),
                 bg=C["bg"], fg=C["muted"]).grid(
                     row=0, column=2, padx=(0, 8), pady=4, sticky="w")
        new_cat = tk.Entry(top, width=20, font=("Arial", 12),
                           bg=C["input"], fg=C["text"], bd=0,
                           insertbackground=C["accent"])
        new_cat.grid(row=0, column=3, padx=(0, 6), pady=4, ipady=4, sticky="ew")

        def add_cat():
            nm = new_cat.get().strip()
            if not nm:
                messagebox.showerror("Error", "Enter name."); return
            if nm in data_ref:
                messagebox.showerror("Error", "Already exists."); return
            try:
                data_ref[nm] = {}
                refresh_cats()
                cat_var.set(nm)
                load_items()
                new_cat.delete(0, tk.END)
            except Exception as e:
                app_log(f"[add_cat] {e}")
                messagebox.showerror("Error", f"Could not add category: {e}")

        def del_cat():
            c = cat_var.get().strip()
            if not c: return
            if messagebox.askyesno("Delete",
                                   f"Delete '{c}' and ALL its items?"):
                try:
                    data_ref.pop(c, None)
                    _save_db(self.svc_data, self.prd_data)
                    refresh_cats()
                    load_items()
                except Exception as e:
                    app_log(f"[del_cat] {e}")
                    messagebox.showerror("Error", f"Could not delete category: {e}")

        ModernButton(top, text="Add", command=add_cat,
                     color=C["teal"], hover_color=C["blue"],
                     width=scaled_value(84, 76, 64), height=scaled_value(30, 30, 26), radius=8,
                     font=("Arial", scaled_value(10, 10, 9), "bold"),
                     ).grid(row=1, column=1, padx=(0, 4), pady=(6, 0), sticky="w")
        ModernButton(top, text="Delete Category", command=del_cat,
                     color=C["red"], hover_color="#c0392b",
                     width=scaled_value(112, 98, 86), height=scaled_value(30, 30, 26), radius=8,
                     font=("Arial", scaled_value(10, 10, 9), "bold"),
                     ).grid(row=1, column=3, padx=(0, 4), pady=(6, 0), sticky="w")

        search_bar = tk.Frame(content, bg=C["bg"])
        search_bar.pack(fill=tk.X, padx=10, pady=(4, 2))
        tk.Label(search_bar, text="Search:", font=("Arial", 11),
                 bg=C["bg"], fg=C["muted"]).pack(side=tk.LEFT, padx=(0, 8))
        list_search_var = tk.StringVar()
        list_search_e = tk.Entry(search_bar, textvariable=list_search_var,
                                 font=("Arial", 11), bg=C["input"], fg=C["text"],
                                 bd=0, insertbackground=C["accent"])
        list_search_e.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)

        product_mode = label == "Products"
        tree_columns = ("name", "price")
        if product_mode:
            tree_columns = ("name", "brand", "category", "variant", "price", "stock")
        self._admin_tree_cols = tree_columns
        tree = ttk.Treeview(content, columns=tree_columns,
                            show="headings", height=12 if product_mode else 14)
        if product_mode:
            tree.heading("name", text="Product Name", anchor="w")
            tree.heading("brand", text="Brand", anchor="w")
            tree.heading("category", text="Category", anchor="w")
            tree.heading("variant", text="Variant", anchor="w")
            tree.heading("price", text="Price (Rs)", anchor="e")
            tree.heading("stock", text="Stock", anchor="center")
            tree.column("name", width=280, anchor="w")
            tree.column("brand", width=140, anchor="w")
            tree.column("category", width=130, anchor="w")
            tree.column("variant", width=110, anchor="w")
            tree.column("price", width=110, anchor="e")
            tree.column("stock", width=90, anchor="center")
        else:
            tree.heading("name",  text="Item Name", anchor="w")
            tree.heading("price", text="Price (Rs)", anchor="e")
            tree.column("name",  width=600, anchor="w")
            tree.column("price", width=160, anchor="e")
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        content.bind("<Configure>", lambda e, tree=tree, product_mode=product_mode: self._resize_admin_tree(tree, product_mode, e.width), add="+")

        frm = tk.Frame(content, bg=C["bg"])
        frm.pack(fill=tk.X, padx=10)
        for col in range(4):
            frm.grid_columnconfigure(col, weight=1 if col in (1, 3) else 0)
        tk.Label(frm, text="Name:", bg=C["bg"],
                 fg=C["muted"]).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        name_e = tk.Entry(frm, width=50, font=("Arial", 12),
                          bg=C["input"], fg=C["text"], bd=0,
                          insertbackground=C["accent"])
        name_e.grid(row=0, column=1, padx=5, pady=5, ipady=4, sticky="ew")
        tk.Label(frm, text="Price:", bg=C["bg"],
                 fg=C["muted"]).grid(row=0, column=2, padx=5, sticky="w")
        price_e = tk.Entry(frm, width=14, font=("Arial", 12),
                           bg=C["input"], fg=C["text"], bd=0,
                           insertbackground=C["accent"])
        price_e.grid(row=0, column=3, padx=5, pady=5, ipady=4, sticky="ew")
        brand_e = pack_e = stock_e = None
        unit_var = tk.StringVar(value="pcs")
        unit_cb = None
        if product_mode:
            tk.Label(frm, text="Brand:", bg=C["bg"],
                     fg=C["muted"]).grid(row=1, column=0, padx=5, pady=5, sticky="w")
            brand_e = tk.Entry(frm, width=20, font=("Arial", 12),
                               bg=C["input"], fg=C["text"], bd=0,
                               insertbackground=C["accent"])
            brand_e.grid(row=1, column=1, padx=5, pady=5, ipady=4, sticky="ew")

            tk.Label(frm, text="Pack Size:", bg=C["bg"],
                     fg=C["muted"]).grid(row=1, column=2, padx=5, pady=5, sticky="w")
            pack_e = tk.Entry(frm, width=14, font=("Arial", 12),
                              bg=C["input"], fg=C["text"], bd=0,
                              insertbackground=C["accent"])
            pack_e.grid(row=1, column=3, padx=5, pady=5, ipady=4, sticky="ew")

            tk.Label(frm, text="Unit:", bg=C["bg"],
                     fg=C["muted"]).grid(row=2, column=0, padx=5, pady=5, sticky="w")
            unit_cb = ttk.Combobox(frm, textvariable=unit_var, state="readonly",
                                   values=["pcs", "ml", "g", "kg", "L"], width=16)
            unit_cb.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

            tk.Label(frm, text="Stock Qty:", bg=C["bg"],
                     fg=C["muted"]).grid(row=2, column=2, padx=5, pady=5, sticky="w")
            stock_e = tk.Entry(frm, width=14, font=("Arial", 12),
                               bg=C["input"], fg=C["text"], bd=0,
                               insertbackground=C["accent"])
            stock_e.insert(0, "0")
            stock_e.grid(row=2, column=3, padx=5, pady=5, ipady=4, sticky="ew")

        # ── Smart Search on name_e ──────────────────
        _ss = {"win": None, "lb": None, "items": []}

        def _adm_typing(e=None, _ss_ref=_ss):
            q   = name_e.get().strip()
            cat = cat_var.get()
            if not q:
                _adm_hide(_ss_ref)
                return
            ql = q.lower()
            if cat and cat != "All":
                pools = [(cat, data_ref.get(cat) or {})]
            else:
                pools = list(data_ref.items())
            matches = []
            for pool_cat, items in pools:
                for nm, pr in items.items():
                    if ql in nm.lower() or ql in pool_cat.lower():
                        matches.append((nm, pr, pool_cat))
            if not matches:
                _adm_hide(_ss_ref)
                return
            _adm_show(_ss_ref, matches)

        def _adm_show(_ss_ref, matches):
            _adm_hide(_ss_ref)
            try:
                x = name_e.winfo_rootx()
                y = name_e.winfo_rooty() + name_e.winfo_height()
                w = name_e.winfo_width()
                win = tk.Toplevel(self.win)
                win.wm_overrideredirect(True)
                fit_toplevel(
                    win, w, min(len(matches) * 26, 180),
                    min_width=max(220, w), min_height=120,
                    resizable=True, anchor="topright",
                    top_offset=y,
                    right_offset=max(20, win.winfo_screenwidth() - (x + w)))
                try:
                    win.transient(self.win.winfo_toplevel())
                    win.lift(self.win.winfo_toplevel())
                except Exception:
                    pass
                _ss_ref["win"] = win
                lb = tk.Listbox(win, font=("Arial", 10),
                                bg=C["card"], fg=C["text"],
                                selectbackground=C["teal"],
                                selectforeground="white",
                                bd=0, activestyle="none")
                lb.pack(fill=tk.BOTH, expand=True)
                _ss_ref["lb"]    = lb
                _ss_ref["items"] = matches
                for nm, pr, pool_cat in matches:
                    lb.insert(tk.END, f"  {nm}  [{pool_cat}]  -  Rs{pr}")
                if lb.size() > 0:
                    lb.selection_set(0)
                    lb.activate(0)

                def _pick(e=None):
                    sel = lb.curselection()
                    if not sel: return
                    nm, pr, pool_cat = matches[sel[0]]
                    name_e.delete(0, tk.END)
                    name_e.insert(0, nm)
                    price_e.delete(0, tk.END)
                    price_e.insert(0, str(pr))
                    if pool_cat:
                        cat_var.set(pool_cat)
                    _adm_hide(_ss_ref)

                lb.bind("<ButtonRelease-1>", _pick)
                lb.bind("<Double-Button-1>", _pick)
                lb.bind("<Return>", _pick)
                lb.bind("<Up>", lambda e: _adm_move(-1, _ss_ref))
                lb.bind("<Down>", lambda e: _adm_move(1, _ss_ref))
                lb.bind("<Escape>", lambda e: _adm_hide(_ss_ref) or "break")
            except Exception:
                pass

        def _adm_hide(_ss_ref=_ss):
            try:
                if _ss_ref.get("win"):
                    _ss_ref["win"].destroy()
                    _ss_ref["win"] = None
                    _ss_ref["lb"]  = None
            except Exception:
                pass

        def _adm_move(delta, _ss_ref=_ss):
            lb = _ss_ref.get("lb")
            if not lb or lb.size() == 0:
                return "break"
            sel = lb.curselection()
            idx = sel[0] if sel else 0
            if sel:
                idx = max(0, min(idx + delta, lb.size() - 1))
            lb.selection_clear(0, tk.END)
            lb.selection_set(idx)
            lb.activate(idx)
            lb.see(idx)
            return "break"

        def _adm_pick_from_popup(e=None, _ss_ref=_ss):
            lb = _ss_ref.get("lb")
            if not lb or lb.size() == 0:
                return "break"
            sel = lb.curselection()
            if not sel:
                lb.selection_set(0)
                lb.activate(0)
                sel = (0,)
            nm, pr, pool_cat = _ss_ref["items"][sel[0]]
            name_e.delete(0, tk.END)
            name_e.insert(0, nm)
            price_e.delete(0, tk.END)
            price_e.insert(0, str(pr))
            if pool_cat:
                cat_var.set(pool_cat)
            _adm_hide(_ss_ref)
            return "break"

        name_e.bind("<KeyRelease>", _adm_typing)
        name_e.bind("<Down>", lambda e: _adm_move(1))
        name_e.bind("<Up>", lambda e: _adm_move(-1))
        name_e.bind("<Return>", _adm_pick_from_popup)
        name_e.bind("<Escape>", lambda e: _adm_hide() or "break")
        name_e.bind("<FocusOut>",
                    lambda e, r=_ss: self.win.after(200, lambda: _adm_hide(r)))

        table_meta = {}

        def refresh_cats():
            cats = set(data_ref.keys())
            if product_mode and use_v5_product_variants_db():
                for row in search_product_variants_v5(""):
                    cat = str(row.get("category_name", "")).strip()
                    if cat:
                        cats.add(cat)
            values = ["All"] + sorted(cats)
            cat_cb["values"] = values
            if values and not cat_var.get():
                cat_var.set("All")

        def load_items():
            try:
                table_meta.clear()
                for i in tree.get_children():
                    tree.delete(i)
                cat = cat_var.get()
                q = list_search_var.get().strip().lower()
                if product_mode and use_v5_product_variants_db():
                    rows = search_product_variants_v5(q)
                    for row in rows:
                        row_cat = str(row.get("category_name", "")).strip()
                        if cat not in ("", "All") and row_cat and row_cat != cat:
                            continue
                        values = (
                            row.get("product_name", ""),
                            row.get("brand_name", ""),
                            row_cat,
                            row.get("pack_label", ""),
                            row.get("sale_price", 0),
                            row.get("stock_qty", 0),
                        )
                        iid = tree.insert("", tk.END, values=values)
                        table_meta[iid] = {"source": "v5", "row": row}
                    legacy_cats = [cat] if cat in data_ref else list(data_ref.keys())
                    if cat in ("", "All"):
                        legacy_cats = list(data_ref.keys())
                    for legacy_cat in legacy_cats:
                        for nm, pr in data_ref.get(legacy_cat, {}).items():
                            if q and q not in nm.lower() and q not in legacy_cat.lower():
                                continue
                            iid = tree.insert("", tk.END, values=(nm, "", legacy_cat, "", pr, 0))
                            table_meta[iid] = {"source": "legacy", "category": legacy_cat, "name": nm}
                    return
                if cat in ("", "All"):
                    cats_to_show = list(data_ref.keys())
                elif cat in data_ref:
                    cats_to_show = [cat]
                else:
                    return
                for current_cat in cats_to_show:
                    for nm, pr in data_ref[current_cat].items():
                        if q and q not in nm.lower() and q not in current_cat.lower():
                            continue
                        iid = tree.insert("", tk.END, values=(nm, pr))
                        table_meta[iid] = {"source": "legacy", "category": current_cat, "name": nm}
            except Exception as e:
                app_log(f"[load_items] {e}")

        cat_cb.bind("<<ComboboxSelected>>",
                    lambda e: load_items())
        list_search_var.trace_add("write", lambda *_: load_items())

        def on_tree_select(e=None):
            sel = tree.selection()
            if not sel: return
            meta = table_meta.get(sel[0], {})
            v = tree.item(sel[0], "values")
            if product_mode and meta.get("source") == "v5":
                row = meta.get("row", {})
                name_e.delete(0, tk.END)
                name_e.insert(0, row.get("product_name", ""))
                price_e.delete(0, tk.END)
                price_e.insert(0, str(row.get("sale_price", 0)))
                if brand_e:
                    brand_e.delete(0, tk.END)
                    brand_e.insert(0, row.get("brand_name", ""))
                if pack_e:
                    pack_e.delete(0, tk.END)
                    pack_e.insert(0, str(row.get("unit_value", "")).rstrip("0").rstrip("."))
                if unit_cb:
                    unit_var.set(row.get("unit_type", "pcs") or "pcs")
                if stock_e:
                    stock_e.delete(0, tk.END)
                    stock_e.insert(0, str(int(float(row.get("stock_qty", 0) or 0))))
                if row.get("category_name"):
                    cat_var.set(row.get("category_name"))
            else:
                name_e.delete(0, tk.END)
                name_e.insert(0, v[0])
                price_e.delete(0, tk.END)
                price_e.insert(0, str(v[1] if not product_mode else v[4]))

        tree.bind("<<TreeviewSelect>>", on_tree_select)

        # Buttons
        bb = tk.Frame(content, bg=C["bg"])
        bb.pack(fill=tk.X, padx=10, pady=6)

        def save_item():
            cat   = cat_var.get()
            nm    = name_e.get().strip()
            pr_s  = price_e.get().strip()
            if not cat or cat == "All":
                messagebox.showerror("Error", "Select a category."); return
            if not nm:
                messagebox.showerror("Error", "Enter item name."); return
            try:
                pr = float(pr_s)
            except ValueError:
                messagebox.showerror("Error", "Enter valid price."); return
            try:
                save_name = nm
                if product_mode:
                    brand = brand_e.get().strip() if brand_e else ""
                    pack_size = pack_e.get().strip() if pack_e else ""
                    unit = unit_var.get().strip() or "pcs"
                    stock_qty = 0
                    try:
                        stock_qty = int((stock_e.get().strip() if stock_e else "0") or "0")
                    except Exception:
                        stock_qty = 0
                    save_name = _product_display_name(nm, brand, pack_size, unit) or nm
                    if use_v5_product_variants_db():
                        create_product_with_variants_v5({
                            "brand_name": brand or "Generic",
                            "category_name": cat,
                            "product_name": nm,
                            "base_name": nm,
                            "variants": [{
                                "variant_name": save_name,
                                "unit_value": pack_size or 1,
                                "unit_type": unit,
                                "pack_label": build_pack_label(pack_size, unit),
                                "bill_label": save_name,
                                "sale_price": pr,
                                "cost_price": pr,
                                "stock_qty": stock_qty,
                                "reorder_level": 0,
                            }],
                        })
                    InventoryService().save_item(
                        _inventory_sync_payload(
                            category=cat,
                            item_name=nm,
                            price=pr,
                            brand=brand,
                            pack_size=pack_size,
                            unit=unit,
                            stock_qty=stock_qty,
                        )
                    )
                data_ref.setdefault(cat, {})[save_name] = pr
                ok = _save_db(self.svc_data, self.prd_data)
                if not ok:
                    raise RuntimeError("save_json returned False")
                from utils import build_item_codes
                build_item_codes(force=True)
                refresh_product_catalog_cache()
                load_items()
                name_e.delete(0, tk.END)
                price_e.delete(0, tk.END)
                if brand_e:
                    brand_e.delete(0, tk.END)
                if pack_e:
                    pack_e.delete(0, tk.END)
                if stock_e:
                    stock_e.delete(0, tk.END); stock_e.insert(0, "0")
                if unit_cb:
                    unit_var.set("pcs")
                messagebox.showinfo("Saved", f"'{save_name}' saved in {cat}.")
            except Exception as e:
                app_log(f"[save_item] {e}")
                messagebox.showerror("Error", f"Could not save item: {e}")

        def delete_item():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Select", "Select an item."); return
            meta = table_meta.get(sel[0], {})
            nm  = tree.item(sel[0], "values")[0]
            cat = meta.get("category", cat_var.get())
            if messagebox.askyesno("Delete", f"Delete '{nm}'?"):
                try:
                    if product_mode and meta.get("source") == "v5" and meta.get("variant_id"):
                        from repositories.product_variants_repo import ProductVariantsRepository
                        ProductVariantsRepository().deactivate_variant(int(meta["variant_id"]))
                    if product_mode:
                        InventoryService().delete_item(nm)
                    data_ref.get(cat, {}).pop(nm, None)
                    _save_db(self.svc_data, self.prd_data)
                    from utils import build_item_codes
                    build_item_codes(force=True)
                    refresh_product_catalog_cache()
                    load_items()
                except Exception as e:
                    app_log(f"[delete_item] {e}")
                    messagebox.showerror("Error", f"Could not delete item: {e}")

        # H9 FIX: Import constants - file size limit and max row limit
        IMPORT_MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB
        IMPORT_MAX_ROWS = 20000  # Prevent DoS via massive files

        def _validate_import_file_size(path: str) -> str | None:
            """Return error message if file exceeds size limit."""
            try:
                size = os.path.getsize(path)
                if size > IMPORT_MAX_FILE_SIZE:
                    return f"File too large ({size // 1024}KB). Maximum allowed: {IMPORT_MAX_FILE_SIZE // 1024}KB."
            except Exception:
                pass
            return None

        def _sanitize_import_string(value: str, field: str = "value", max_len: int = 200) -> str:
            """Sanitize a single import field value.

            H9 FIX: Strip dangerous characters, limit length, and prevent
            injection through field values in imported data.
            """
            value = str(value or "").strip()
            # Limit length to prevent memory exhaustion and overflow
            if len(value) > max_len:
                app_log(f"[import] {field} truncated from {len(value)} to {max_len}: {value[:50]}...")
                value = value[:max_len]
            # Strip null bytes and control characters (except common whitespace)
            value = value.replace("\x00", "").replace("\r", "")
            # Remove potentially dangerous HTML/script content in names
            if "<" in value or ">" in value:
                value = value.replace("<", "").replace(">", "")
            return value

        def import_json_items():
            path = filedialog.askopenfilename(
                title="Import JSON",
                filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")],
            )
            if not path:
                return

            # H9 FIX: Validate file size before reading
            size_err = _validate_import_file_size(path)
            if size_err:
                messagebox.showerror("Import Error", size_err)
                return

            # Validate file extension
            if not path.lower().endswith(".json"):
                messagebox.showerror("Import Error", "Only JSON files are allowed.")
                return

            try:
                with open(path, "r", encoding="utf-8") as fh:
                    payload = json.load(fh)
                rows = _normalize_import_rows(payload)

                # H9 FIX: Limit number of rows to prevent DoS
                if len(rows) > IMPORT_MAX_ROWS:
                    messagebox.showerror(
                        "Import Error",
                        f"Too many rows ({len(rows)}). Maximum allowed: {IMPORT_MAX_ROWS}."
                    )
                    return

                summary = {"rows_read": len(rows), "imported": 0, "created": 0, "updated": 0,
                           "skipped": 0, "errors": 0, "error_lines": []}
                for idx, row in enumerate(rows, start=1):
                    try:
                        # H9 FIX: Sanitize all import fields
                        item_name = _sanitize_import_string(
                            row.get("Item Name") or row.get("name") or "", "item_name")
                        category = _sanitize_import_string(
                            row.get("Category") or row.get("category") or cat_var.get() or "", "category")
                        if not item_name or not category:
                            summary["skipped"] += 1
                            summary["error_lines"].append(f"Row {idx}: missing item name or category")
                            continue
                        brand = _sanitize_import_string(
                            row.get("Brand") or row.get("brand") or "", "brand", max_len=100)
                        pack_size = _sanitize_import_string(
                            row.get("Pack Size") or row.get("pack_size") or "", "pack_size", max_len=50)
                        unit = _sanitize_import_string(
                            row.get("Unit") or row.get("unit") or "pcs", "unit", max_len=20) or "pcs"

                        # H9 FIX: Validate numeric fields before conversion
                        try:
                            price = float(row.get("Price") or row.get("price") or 0)
                            if price < 0:
                                price = 0.0
                            stock = int(float(row.get("Stock") or row.get("stock") or 0))
                            if stock < 0:
                                stock = 0
                        except (ValueError, TypeError) as ve:
                            summary["skipped"] += 1
                            summary["error_lines"].append(f"Row {idx}: invalid numeric value: {ve}")
                            continue
                        save_name = _product_display_name(item_name, brand, pack_size, unit) if product_mode else item_name
                        existing = save_name in data_ref.get(category, {})
                        data_ref.setdefault(category, {})[save_name] = price
                        if product_mode and use_v5_product_variants_db():
                            create_product_with_variants_v5({
                                "brand_name": brand or "Generic",
                                "category_name": category,
                                "product_name": item_name,
                                "base_name": item_name,
                                "variants": [{
                                    "variant_name": save_name,
                                    "unit_value": pack_size or 1,
                                    "unit_type": unit,
                                    "pack_label": build_pack_label(pack_size, unit),
                                    "bill_label": save_name,
                                    "sale_price": price,
                                    "cost_price": price,
                                    "stock_qty": stock,
                                    "reorder_level": 0,
                                }],
                            })
                            InventoryService().save_item(
                                _inventory_sync_payload(
                                    category=category,
                                    item_name=item_name,
                                    price=price,
                                    brand=brand,
                                    pack_size=pack_size,
                                    unit=unit,
                                    stock_qty=stock,
                                )
                            )
                        summary["imported"] += 1
                        if existing:
                            summary["updated"] += 1
                        else:
                            summary["created"] += 1
                    except Exception as row_error:
                        summary["errors"] += 1
                        summary["error_lines"].append(f"Row {idx}: {row_error}")
                _save_db(self.svc_data, self.prd_data)
                from utils import build_item_codes
                build_item_codes(force=True)
                refresh_product_catalog_cache()
                refresh_cats()
                load_items()
                _show_import_summary("Import Complete", path, summary)
            except Exception as e:
                app_log(f"[import_json_items] {e}")
                messagebox.showerror("Import Error", f"Could not import JSON: {e}")

        def import_excel_items():
            path = filedialog.askopenfilename(
                title="Import Excel",
                filetypes=[("Excel Files", "*.xlsx;*.xlsm"), ("All Files", "*.*")],
            )
            if not path:
                return

            # H9 FIX: Validate file size
            size_err = _validate_import_file_size(path)
            if size_err:
                messagebox.showerror("Import Error", size_err)
                return

            # Validate file extension
            if not (path.lower().endswith(".xlsx") or path.lower().endswith(".xlsm")):
                messagebox.showerror("Import Error", "Only Excel files (.xlsx, .xlsm) are allowed.")
                return

            try:
                from openpyxl import load_workbook
            except Exception:
                messagebox.showerror("Import Error", "openpyxl is not installed in this build.")
                return
            try:
                wb = load_workbook(path, data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                header_map = {h.lower(): idx for idx, h in enumerate(headers)}
                total_rows = max(ws.max_row - 1, 0)

                # H9 FIX: Limit rows to prevent DoS
                if total_rows > IMPORT_MAX_ROWS:
                    messagebox.showerror(
                        "Import Error",
                        f"Too many rows ({total_rows}). Maximum allowed: {IMPORT_MAX_ROWS}."
                    )
                    return

                summary = {"rows_read": total_rows, "imported": 0, "created": 0, "updated": 0,
                           "skipped": 0, "errors": 0, "error_lines": []}
                for excel_row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        excel_row_no = excel_row_no  # noqa -- row number used below
                        # H9 FIX: Sanitize all import fields
                        item_name = _sanitize_import_string(
                            row[header_map.get("item name", -1)] if "item name" in header_map else ""
                        ) if "item name" in header_map else ""
                        category = _sanitize_import_string(
                            row[header_map.get("category", -1)] if "category" in header_map else ""
                        ) if "category" in header_map else ""
                        if not item_name or not category:
                            summary["skipped"] += 1
                            summary["error_lines"].append(f"Row {excel_row_no}: missing item name or category")
                            continue
                        brand = _sanitize_import_string(
                            row[header_map.get("brand", -1)] if "brand" in header_map else "", "brand", max_len=100
                        ) if "brand" in header_map else ""
                        pack_size = _sanitize_import_string(
                            row[header_map.get("pack size", -1)] if "pack size" in header_map else "", "pack_size", max_len=50
                        ) if "pack size" in header_map else ""
                        unit = _sanitize_import_string(
                            row[header_map.get("unit", -1)] if "unit" in header_map else "pcs", "unit", max_len=20
                        ) if "unit" in header_map else "pcs" or "pcs"

                        # H9 FIX: Validate numeric fields
                        try:
                            price = float(row[header_map.get("price", -1)] or 0) if "price" in header_map else 0.0
                            if price < 0:
                                price = 0.0
                            stock = int(float(row[header_map.get("stock", -1)] or 0)) if "stock" in header_map else 0
                            if stock < 0:
                                stock = 0
                        except (ValueError, TypeError) as ve:
                            summary["skipped"] += 1
                            summary["error_lines"].append(f"Row {excel_row_no}: invalid numeric value: {ve}")
                            continue
                        save_name = _product_display_name(item_name, brand, pack_size, unit) if product_mode else item_name
                        existing = save_name in data_ref.get(category, {})
                        data_ref.setdefault(category, {})[save_name] = price
                        if product_mode and use_v5_product_variants_db():
                            create_product_with_variants_v5({
                                "brand_name": brand or "Generic",
                                "category_name": category,
                                "product_name": item_name,
                                "base_name": item_name,
                                "variants": [{
                                    "variant_name": save_name,
                                    "unit_value": pack_size or 1,
                                    "unit_type": unit,
                                    "pack_label": build_pack_label(pack_size, unit),
                                    "bill_label": save_name,
                                    "sale_price": price,
                                    "cost_price": price,
                                    "stock_qty": stock,
                                    "reorder_level": 0,
                                }],
                            })
                            InventoryService().save_item(
                                _inventory_sync_payload(
                                    category=category,
                                    item_name=item_name,
                                    price=price,
                                    brand=brand,
                                    pack_size=pack_size,
                                    unit=unit,
                                    stock_qty=stock,
                                )
                            )
                        summary["imported"] += 1
                        if existing:
                            summary["updated"] += 1
                        else:
                            summary["created"] += 1
                    except Exception as row_error:
                        summary["errors"] += 1
                        summary["error_lines"].append(f"Row {excel_row_no}: {row_error}")
                _save_db(self.svc_data, self.prd_data)
                from utils import build_item_codes
                build_item_codes(force=True)
                refresh_product_catalog_cache()
                refresh_cats()
                load_items()
                _show_import_summary("Import Complete", path, summary)
            except Exception as e:
                app_log(f"[import_excel_items] {e}")
                messagebox.showerror("Import Error", f"Could not import Excel: {e}")

        save_icon = get_action_icon("save")
        delete_icon = get_action_icon("delete")
        refresh_icon = get_action_icon("refresh")
        json_icon = get_action_icon("import_json")
        excel_icon = get_action_icon("import_excel")

        for txt, clr, hclr, cmd in [
            ("Save / Update", C["teal"], C["blue"], save_item),
            ("Delete Item",   C["red"],  "#c0392b", delete_item),
        ]:
            btn_icon = save_icon if txt == "Save / Update" else delete_icon
            ModernButton(bb, text=txt, image=btn_icon, compound="left", command=cmd,
                         color=clr, hover_color=hclr,
                         width=scaled_value(150, 132, 112), height=scaled_value(34, 32, 28), radius=8,
                          font=("Arial", scaled_value(10, 10, 9), "bold"),
                          ).pack(side=tk.LEFT, padx=3)
        if product_mode:
            ModernButton(bb, text="Refresh", image=refresh_icon, compound="left", command=load_items,
                         color=C["blue"], hover_color="#154360",
                         width=scaled_value(120, 110, 96), height=scaled_value(34, 32, 28), radius=8,
                         font=("Arial", scaled_value(10, 10, 9), "bold")).pack(side=tk.LEFT, padx=3)
            ModernButton(bb, text="Import JSON", image=json_icon, compound="left", command=import_json_items,
                         color=C["purple"], hover_color="#6c4ccf",
                         width=scaled_value(132, 118, 100), height=scaled_value(34, 32, 28), radius=8,
                         font=("Arial", scaled_value(10, 10, 9), "bold")).pack(side=tk.LEFT, padx=3)
            ModernButton(bb, text="Import Excel", image=excel_icon, compound="left", command=import_excel_items,
                         color=C["orange"], hover_color="#d97a00",
                         width=scaled_value(132, 118, 100), height=scaled_value(34, 32, 28), radius=8,
                         font=("Arial", scaled_value(10, 10, 9), "bold")).pack(side=tk.LEFT, padx=3)

        refresh_cats()
        load_items()

    # ─────────── Users tab ────────────
    def _build_users_tab(self, parent):
        # Bug 12 fix: use self.current_user (now properly initialized)
        def _open_user_mgr():
            try:
                UserManagerWindow(self.win, self.current_user)
            except Exception as e:
                app_log(f"[_build_users_tab] open user manager: {e}")
                messagebox.showerror("Error", f"Could not open User Management: {e}")

        try:
            ModernButton(parent, text="Open User Management",
                         command=_open_user_mgr,
                         color=C["teal"], hover_color=C["blue"],
                         width=scaled_value(220, 200, 168), height=scaled_value(40, 36, 32), radius=8,
                         font=("Arial", scaled_value(11, 10, 9), "bold"),
                         ).pack(pady=30)

            tk.Label(parent,
                     text="Add, edit, reset passwords, and manage staff login accounts.",
                     bg=C["bg"], fg=C["muted"],
                     font=("Arial", 11)).pack()
        except Exception as e:
            app_log(f"[_build_users_tab] {e}")

    def _resize_admin_tree(self, tree, product_mode, width):
        width = max(520, width - 36)
        if product_mode:
            col_map = {
                "name": max(180, int(width * 0.28)),
                "brand": max(110, int(width * 0.15)),
                "category": max(110, int(width * 0.15)),
                "variant": max(96, int(width * 0.13)),
                "price": max(96, int(width * 0.12)),
            }
            used = sum(col_map.values())
            col_map["stock"] = max(82, width - used)
        else:
            col_map = {
                "name": max(260, int(width * 0.72)),
                "price": max(120, width - max(260, int(width * 0.72))),
            }
        for col, val in col_map.items():
            tree.column(col, width=val)

    # ─────────── Backup & Logs tab ────────────
    def _build_backup_tab(self, parent):
        """V5.6.1 Phase 1 — Backup & Activity quick-access."""
        from scheduled_backup import get_scheduled_config, run_scheduled_backup
        from backup_system import sync_offline_backup, normalize_backup_folder, get_backup_config

        content, _canvas, _container = make_scrollable(parent, bg=C["bg"], padx=0, pady=0)
        f = tk.Frame(content, bg=C["bg"], padx=16, pady=12)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="Backup & Activity", font=("Arial", 14, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(anchor="w", pady=(0, 6))
        tk.Frame(f, bg=C["teal"], height=2).pack(fill=tk.X, pady=(0, 12))

        # Manual backup button
        def _manual_backup():
            cfg = get_backup_config()
            folder = cfg.get("folder", "").strip()
            if not folder:
                from tkinter import filedialog
                folder = filedialog.askdirectory(
                    title="Select Backup Folder", parent=self.win)
                if not folder:
                    return
                cfg["folder"] = folder
                from backup_system import save_backup_config
                save_backup_config(cfg)
            from tkinter import messagebox as _mb
            try:
                ok, errs, root = run_scheduled_backup()
                if ok > 0:
                    _mb.showinfo("Backup Done",
                                 f"Backed up {ok} items.\nFolder: {root}")
                else:
                    _mb.showwarning("Backup Failed", "\n".join(errs[:5]))
            except Exception as e:
                _mb.showerror("Error", f"Backup failed: {e}")

        ModernButton(f, text="Run Backup Now", command=_manual_backup,
                     color=C["teal"], hover_color=C["blue"],
                     width=scaled_value(200, 180, 160), height=36, radius=8,
                     font=("Arial", scaled_value(11, 10, 9), "bold"),
                     ).pack(anchor="w", pady=(0, 8))

        # Status labels
        sched_cfg = get_scheduled_config()
        base_cfg = get_backup_config()
        if sched_cfg.get("enabled"):
            status_text = f"Scheduled: {sched_cfg.get('frequency', 'daily')} at {sched_cfg.get('time', '02:00')}"
        else:
            status_text = "Scheduled backup: disabled"
        last = base_cfg.get("last_backup", "Never")
        last_sched = sched_cfg.get("last_backup", "Never")
        last_err = sched_cfg.get("last_error", "")

        for txt, color_val in [
            (f"Last manual backup: {last}", C["muted"]),
            (f"Last scheduled backup: {last_sched}", C["muted"]),
            (status_text, C["teal"] if sched_cfg.get("enabled") else C["orange"]),
        ]:
            tk.Label(f, text=txt, bg=C["bg"], fg=color_val,
                     font=("Arial", 10)).pack(anchor="w", pady=1)
        if last_err:
            tk.Label(f, text=f"Last error: {last_err}", bg=C["bg"],
                     fg=C["red"], font=("Arial", 9)).pack(anchor="w", pady=1)

        tk.Frame(f, bg=C["sidebar"], height=1).pack(fill=tk.X, pady=(10, 10))

        # Activity Log viewer
        def _open_log():
            try:
                from activity_log import get_event_count
                count = get_event_count()
            except Exception:
                count = 0
            try:
                from activity_log import show_activity_log_viewer
                show_activity_log_viewer(self.win)
            except Exception as e:
                messagebox.showerror("Error",
                                      f"Could not open activity log:\n{e}")

        ModernButton(f, text="View Activity Log", command=_open_log,
                     color=C["blue"], hover_color="#154360",
                     width=scaled_value(200, 180, 160), height=36, radius=8,
                     font=("Arial", scaled_value(11, 10, 9), "bold"),
                     ).pack(anchor="w", pady=(0, 6))

        try:
            from activity_log import get_event_count
            cnt = get_event_count()
        except Exception:
            cnt = 0
        tk.Label(f, text=f"Total activity entries: {cnt}", bg=C["bg"], fg=C["muted"],
                 font=("Arial", 10)).pack(anchor="w", pady=(0, 8))


