"""Report export helpers extracted from reports.py."""
from __future__ import annotations

import csv
import os

from utils import DATA_DIR, F_REPORT, today_str


def export_current_tree_csv(path: str, tree) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        try:
            cols = [tree.heading(c)["text"] for c in tree["columns"]]
            writer.writerow(cols)
            for row_id in tree.get_children():
                writer.writerow(tree.item(row_id, "values"))
        except Exception:
            if os.path.exists(F_REPORT):
                with open(F_REPORT, "r", encoding="utf-8") as src:
                    for line in src:
                        writer.writerow(line.strip().split(","))


def export_report_excel_or_csv(rows: list) -> tuple[str, bool]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1a1a2e")
        headers = ["Date", "Invoice", "Customer", "Phone", "Payment", "Discount", "Total", "Items"]
        for ci, heading in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=heading)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(rows, 2):
            items_readable = " | ".join(
                f"{seg.split('~')[1]} x{seg.split('~')[3]}"
                for seg in row.get("items_raw", "").split("|")
                if len(seg.split("~")) == 4
            ) if row.get("items_raw") else ""
            for ci, value in enumerate([
                row.get("date", ""),
                row.get("invoice", ""),
                row.get("name", ""),
                row.get("phone", ""),
                row.get("payment", ""),
                row.get("discount", 0),
                row.get("total", 0),
                items_readable,
            ], 1):
                ws.cell(row=ri, column=ci, value=value)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        out = os.path.join(DATA_DIR, f"sales_export_{today_str()}.xlsx")
        wb.save(out)
        return out, True
    except ImportError:
        out = os.path.join(DATA_DIR, f"sales_export_{today_str()}.csv")
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Date", "Invoice", "Customer", "Phone", "Payment", "Discount", "Total"])
            for row in rows:
                writer.writerow([
                    row.get("date", ""),
                    row.get("invoice", ""),
                    row.get("name", ""),
                    row.get("phone", ""),
                    row.get("payment", ""),
                    row.get("discount", 0),
                    row.get("total", 0),
                ])
        return out, False
